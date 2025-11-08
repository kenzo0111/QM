from __future__ import annotations

import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from dataclasses import dataclass, field
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
import random
import os
import json
from pathlib import Path

#
# --- New imports for map visualization ---
#
import folium
import osmnx as ox
import geopandas as gpd
from shapely.geometry import Point, LineString, MultiLineString
from branca.element import MacroElement, Template

# Vehicle types with mass ranges (kg)
vehicle_types = {
    "Car": {"mass_range": (1200, 1500)},
    "Truck": {"mass_range": (5000, 10000)},
    "Motorcycle": {"mass_range": (150, 300)},
    "Bus": {"mass_range": (8000, 12000)}
}

# Accident types
accident_types = {
    "rear-end": {"description": "Rear-end collision"},
    "head-on": {"description": "Head-on collision"},
    "side-impact": {"description": "Side-impact (T-bone) collision"},
    "pedestrian": {"description": "Vehicle-pedestrian collision"}
}

# Possible causes
causes = ["Overspeeding", "Drunk Driving", "Mechanical Failure", "Fatigue", "Poor Visibility", "Reckless Driving"]

# Road and lighting conditions
road_conditions = ["dry", "wet", "slippery"]
lighting_conditions = ["good", "poor"]

# Possible recommendations based on accident factors
recommendations_db = {
    "Overspeeding": ["Implement speed cameras", "Increase speed limit enforcement", "Add speed bumps in high-risk areas"],
    "Drunk Driving": ["Stricter DUI checkpoints", "Public awareness campaigns on drinking and driving", "Install breathalyzer devices"],
    "Mechanical Failure": ["Mandatory vehicle inspections", "Improve maintenance education", "Create emergency roadside assistance programs"],
    "Fatigue": ["Install rest areas along highways", "Campaigns on driver fatigue awareness", "Regulate maximum driving hours"],
    "Poor Visibility": ["Upgrade street lighting", "Install reflective road signs", "Improve weather-responsive lighting systems"],
    "Reckless Driving": ["Enhanced driver training programs", "Increase traffic police presence", "Implement point-based license system"],
    "Human Error": ["Driver education and awareness programs", "Improve road signage", "Traffic signal optimization"],
    "dry": ["Regular road maintenance", "Pothole repair programs"],
    "wet": ["Improve drainage systems", "Install anti-skid road surfaces", "Weather-based speed limit adjustments"],
    "slippery": ["Apply anti-slip treatments", "Install warning signs for slippery areas", "Tire tread depth checks"],
    "good": ["Maintain current lighting standards"],
    "poor": ["Install additional street lights", "Upgrade to LED lighting systems", "Solar-powered lighting solutions"],
    "rear-end": ["Increase following distance education", "Install rear-end collision warning systems"],
    "head-on": ["Install median barriers", "Improve lane markings", "One-way traffic in narrow roads"],
    "side-impact": ["Improve intersection visibility", "Install turning signal cameras", "Enhanced crosswalk protections"],
    "pedestrian": ["Install pedestrian crossings", "Speed reduction near schools/zones", "Pedestrian safety education"]
}

# Human factor profiles influence driver reaction and braking effectiveness
human_factor_profiles = {
    "None": {"reaction_multiplier": 1.0, "braking_multiplier": 1.0, "risk_multiplier": 1.0, "notes": "Driver operating within normal alertness."},
    "None listed": {"reaction_multiplier": 1.0, "braking_multiplier": 1.0, "risk_multiplier": 1.0, "notes": "No reported impairment."},
    "Under influence": {"reaction_multiplier": 1.6, "braking_multiplier": 0.85, "risk_multiplier": 1.5, "notes": "Impaired decision making due to alcohol or substances."},
    "Drunk Driving": {"reaction_multiplier": 1.7, "braking_multiplier": 0.85, "risk_multiplier": 1.6, "notes": "Blood alcohol levels hamper perception and coordination."},
    "Fatigue": {"reaction_multiplier": 1.4, "braking_multiplier": 0.9, "risk_multiplier": 1.3, "notes": "Delayed reaction because of drowsiness."},
    "Sleepy driver": {"reaction_multiplier": 1.5, "braking_multiplier": 0.9, "risk_multiplier": 1.4, "notes": "Driver vigilance reduced due to lack of rest."},
    "None reported": {"reaction_multiplier": 1.0, "braking_multiplier": 1.0, "risk_multiplier": 1.0, "notes": "No recorded impairment."}
}

# Road location profiles capture geometry characteristics gathered from field observations
location_profiles = {
    "Bayabas": {"slope": 4.5, "curvature": "blind_curve", "default_weather": "Rainy", "recommended_interventions": ["road_surface_treatment", "improved_street_lighting"]},
    "Main Highway": {"slope": 1.0, "curvature": "straight", "default_weather": "Sunny", "recommended_interventions": ["speed_checkpoint", "driver_safety_campaign"]},
    "Barangay 1": {"slope": 2.0, "curvature": "slight_curve", "default_weather": "Cloudy", "recommended_interventions": ["road_surface_treatment"]},
    "Barangay 2": {"slope": 1.5, "curvature": "intersection", "default_weather": "Rainy", "recommended_interventions": ["improved_street_lighting", "community_seminar"]},
    "Barangay 3": {"slope": 3.0, "curvature": "blind_curve", "default_weather": "Cloudy", "recommended_interventions": ["road_surface_treatment", "speed_checkpoint"]},
    "Barangay 4": {"slope": 2.5, "curvature": "intersection", "default_weather": "Rainy", "recommended_interventions": ["improved_street_lighting"]},
    "Barangay 5": {"slope": 3.5, "curvature": "blind_curve", "default_weather": "Rainy", "recommended_interventions": ["road_surface_treatment", "community_seminar"]}
}

# Intervention effects align with Materials & Methods (e.g., checkpoints, seminars, lighting upgrades)
intervention_effects = {
    "improved_street_lighting": {"lighting_override": "good", "reaction_multiplier": 0.75, "description": "Installation of additional LED street lights."},
    "road_surface_treatment": {"friction_bonus": 0.15, "road_condition_override": "dry", "description": "Application of anti-skid and pothole repairs."},
    "speed_checkpoint": {"speed_reduction": 0.25, "driver_risk_multiplier": 0.9, "description": "Random enforcement checkpoints to deter overspeeding."},
    "driver_safety_campaign": {"reaction_multiplier": 0.9, "driver_risk_multiplier": 0.95, "description": "Educational programs on defensive driving and Systems Theory interactions."},
    "community_seminar": {"reaction_multiplier": 0.95, "driver_risk_multiplier": 0.95, "description": "Barangay-wide awareness seminars targeting human factors."}
}


@dataclass
class DriverProfile:
    """Represents the human factor component affecting reaction and braking."""

    factor: str = "None"
    reaction_multiplier: float = 1.0
    braking_multiplier: float = 1.0
    risk_multiplier: float = 1.0
    notes: str = ""

    @classmethod
    def from_factor(cls, factor: str) -> "DriverProfile":
        normalized = (factor or "None").strip()
        profile_key = None
        for key in human_factor_profiles:
            if normalized.lower() == key.lower():
                profile_key = key
                break
        if profile_key is None:
            profile_key = "None"
        profile_data = human_factor_profiles.get(profile_key, human_factor_profiles["None"])
        return cls(
            factor=profile_key,
            reaction_multiplier=profile_data["reaction_multiplier"],
            braking_multiplier=profile_data["braking_multiplier"],
            risk_multiplier=profile_data["risk_multiplier"],
            notes=profile_data["notes"]
        )


@dataclass
class EnvironmentState:
    """Encodes the road, lighting, and weather conditions for the simulation."""

    road_condition: str
    lighting_condition: str
    weather_condition: str
    slope: float = 0.0  # degrees incline/decline approximation
    curvature: str = "straight"  # straight, intersection, blind_curve, etc.

    def get_effective_friction(self, driver_profile: DriverProfile | None = None) -> float:
        base_friction = {"dry": 0.82, "wet": 0.55, "slippery": 0.18}.get(self.road_condition, 0.75)
        # Wetter weather decreases traction further
        if self.weather_condition.lower() in {"rainy", "storm", "typhoon"}:
            base_friction *= 0.9
        # Uphill/Downhill adjustments based on Systems Theory interaction of vehicle-road
        base_friction *= max(0.7, 1 - abs(self.slope) * 0.01)
        if self.curvature in {"blind_curve", "sharp_turn"}:
            base_friction *= 0.95
        if driver_profile:
            base_friction *= driver_profile.braking_multiplier
        for max_limit in (1.2,):
            base_friction = min(base_friction, max_limit)
        return max(base_friction, 0.05)

    def get_effective_reaction(self, driver_profile: DriverProfile | None = None) -> float:
        base_reaction = {"good": 0.55, "poor": 1.5}.get(self.lighting_condition, 0.8)
        if self.weather_condition.lower() in {"rainy", "foggy", "storm"}:
            base_reaction *= 1.2
        if driver_profile:
            base_reaction *= driver_profile.reaction_multiplier
        return max(0.3, min(base_reaction, 3.5))

    def describe(self) -> str:
        return (
            f"Road: {self.road_condition}, Lighting: {self.lighting_condition}, Weather: {self.weather_condition}, "
            f"Slope: {self.slope:.1f}°, Geometry: {self.curvature}"
        )


@dataclass
class InterventionPlan:
    """Represents selected interventions for what-if analysis."""

    interventions: list[str] = field(default_factory=list)

    def describe(self) -> str:
        if not self.interventions:
            return "No interventions applied."
        return ", ".join(
            f"{name.replace('_', ' ').title()}" for name in self.interventions
        )


def build_environment_state(location: str, road_condition: str, lighting_condition: str, weather_condition: str) -> EnvironmentState:
    """Create an EnvironmentState incorporating field observations per location."""

    profile = location_profiles.get(location, {})
    slope = profile.get("slope", 0.0)
    curvature = profile.get("curvature", "straight")
    # If weather not provided, fallback to location default gleaned from field observations
    weather = weather_condition or profile.get("default_weather", "Sunny")
    return EnvironmentState(
        road_condition=road_condition,
        lighting_condition=lighting_condition,
        weather_condition=weather,
        slope=slope,
        curvature=curvature
    )


def suggest_interventions(location: str, cause: str, road_condition: str, lighting_condition: str, human_factor: str) -> InterventionPlan:
    """Suggest interventions grounded in Materials & Methods and Systems Theory."""

    plan = InterventionPlan()
    profile = location_profiles.get(location, {})
    plan.interventions.extend(profile.get("recommended_interventions", []))

    if lighting_condition == "poor" and "improved_street_lighting" not in plan.interventions:
        plan.interventions.append("improved_street_lighting")
    if road_condition in {"wet", "slippery"} and "road_surface_treatment" not in plan.interventions:
        plan.interventions.append("road_surface_treatment")
    if cause == "Overspeeding" and "speed_checkpoint" not in plan.interventions:
        plan.interventions.append("speed_checkpoint")
    if human_factor.lower() in {"under influence", "drunk driving"} and "speed_checkpoint" not in plan.interventions:
        plan.interventions.append("speed_checkpoint")
    if human_factor.lower() in {"fatigue", "sleepy driver"} and "driver_safety_campaign" not in plan.interventions:
        plan.interventions.append("driver_safety_campaign")

    # Ensure uniqueness while preserving order
    seen = set()
    deduped = []
    for intervention in plan.interventions:
        if intervention not in seen:
            deduped.append(intervention)
            seen.add(intervention)
    plan.interventions = deduped
    return plan


def calculate_risk_score(impact_force: float, severity: str, environment: EnvironmentState | None, driver_profile: DriverProfile | None) -> float:
    """Compute a 0-10 risk score blending vehicle, road, and human factors."""

    severity_weights = {"none": 0.0, "minor": 4.0, "moderate": 7.0, "severe": 9.0}
    base_score = severity_weights.get(severity.lower(), 5.0)
    base_score += min(impact_force / 150000, 2.5)

    if environment:
        if environment.road_condition != "dry":
            base_score += 0.7
        if environment.lighting_condition == "poor":
            base_score += 0.6
        base_score += min(abs(environment.slope) / 10.0, 1.0)
        if environment.curvature in {"blind_curve", "intersection"}:
            base_score += 0.5

    if driver_profile:
        base_score *= driver_profile.risk_multiplier

    return round(max(0.0, min(10.0, base_score)), 2)


def generate_systemic_summary(cause: str, driver_profile: DriverProfile | None, environment: EnvironmentState | None, plan: InterventionPlan | None) -> str:
    """Create a short paragraph tying drivers, vehicles, road, and environment (Systems Theory)."""

    components = []
    if driver_profile:
        components.append(f"Driver state ({driver_profile.factor}) affected reaction time ({driver_profile.notes})")
    if environment:
        components.append(f"Road geometry ({environment.curvature}) with {environment.road_condition} surface influenced braking distance")
        if environment.lighting_condition == "poor":
            components.append("Limited lighting stretched driver perception")
    components.append(f"Primary cause flagged: {cause}")
    if plan and plan.interventions:
        named = ", ".join(intervention.replace('_', ' ') for intervention in plan.interventions)
        components.append(f"Interventions tested: {named}")
    return "; ".join(components) + "."


def validate_with_historical(location: str, severity: str) -> str:
    """Compare simulated severity against historical distribution for validation."""

    try:
        historical = pd.read_csv('accident_data.csv')
    except Exception:
        return "Historical validation unavailable (no data)."

    if historical.empty or 'severity' not in historical.columns:
        return "Historical validation unavailable (insufficient data)."

    location_data = historical[historical['location'] == location]
    if location_data.empty:
        return f"No recorded accidents in {location} for validation."

    severity_counts = location_data['severity'].astype(str).str.lower().value_counts()
    total = severity_counts.sum()
    match_count = severity_counts.get(severity.lower(), 0)
    percentage = (match_count / total) * 100 if total else 0
    return f"{percentage:.1f}% of historical accidents in {location} share {severity.lower()} severity."


def determine_angle(accident_type: str) -> int:
    angle_map = {"rear-end": 0, "head-on": 180, "side-impact": 90, "pedestrian": 15}
    return angle_map.get(accident_type, 0)


def prepare_accident_entities(accident_type: str, primary_vehicle_type: str) -> tuple[dict | None, dict | None, dict | None, int]:
    """Create baseline entity specifications for repeatable simulation/what-if runs."""

    def vehicle_spec(position: float, name_suffix: str, velocity: float | None = None, override_type: str | None = None) -> dict:
        veh_type = override_type if override_type else primary_vehicle_type
        mass = random.randint(*vehicle_types[veh_type]["mass_range"])
        speed = velocity if velocity is not None else random.uniform(15, 35)
        return {"mass": mass, "velocity": speed, "position": position, "name": f"{veh_type} {name_suffix}"}

    vehicle1 = vehicle_spec(0, "1")
    vehicle2 = None
    pedestrian = None

    if accident_type == "rear-end":
        vehicle2 = vehicle_spec(50, "2")
        if vehicle1["velocity"] <= vehicle2["velocity"]:
            vehicle1["velocity"] = vehicle2["velocity"] + random.uniform(5, 10)
    elif accident_type == "head-on":
        vehicle1["velocity"] = random.uniform(15, 25)
        vehicle2 = vehicle_spec(100, "2", velocity=-random.uniform(15, 25))
    elif accident_type == "side-impact":
        vehicle1["velocity"] = random.uniform(15, 25)
        vehicle2 = vehicle_spec(50, "2", velocity=0)
    elif accident_type == "pedestrian":
        vehicle1["velocity"] = random.uniform(15, 35)
        pedestrian = {"mass": random.randint(50, 90), "position": random.uniform(40, 60), "name": "Pedestrian"}

    angle = determine_angle(accident_type)
    return vehicle1, vehicle2, pedestrian, angle


def instantiate_entities(vehicle1_spec: dict | None, vehicle2_spec: dict | None, pedestrian_spec: dict | None) -> tuple[Vehicle | None, Vehicle | None, Pedestrian | None]:
    vehicle1 = Vehicle(vehicle1_spec["mass"], vehicle1_spec["velocity"], position=vehicle1_spec["position"], name=vehicle1_spec["name"]) if vehicle1_spec else None
    vehicle2 = Vehicle(vehicle2_spec["mass"], vehicle2_spec["velocity"], position=vehicle2_spec["position"], name=vehicle2_spec["name"]) if vehicle2_spec else None
    pedestrian = Pedestrian(mass=pedestrian_spec["mass"], position=pedestrian_spec["position"], name=pedestrian_spec["name"]) if pedestrian_spec else None
    return vehicle1, vehicle2, pedestrian


def run_monte_carlo_simulations(runs: int = 50, apply_interventions: bool = False) -> dict:
    """Execute multiple simulations to observe severity distribution trends."""

    results = []
    for _ in range(runs):
        cause, location, primary_vehicle_type, road_condition, lighting_condition, weather_condition, human_factor = sample_from_data_sources()
        environment = build_environment_state(location, road_condition, lighting_condition, weather_condition)
        driver_profile = DriverProfile.from_factor(human_factor)
        intervention_plan = suggest_interventions(location, cause, environment.road_condition, environment.lighting_condition, human_factor)
        accident_type = random.choice(list(accident_types.keys()))
        vehicle1_spec, vehicle2_spec, pedestrian_spec, angle = prepare_accident_entities(accident_type, primary_vehicle_type)
        vehicle1, vehicle2, pedestrian = instantiate_entities(vehicle1_spec, vehicle2_spec, pedestrian_spec)

        _, _, _, _, _, impact_force, severity, risk_score, _ = simulate_collision(
            vehicle1,
            vehicle2,
            pedestrian=pedestrian,
            accident_type=accident_type,
            road_condition=environment.road_condition,
            lighting_condition=environment.lighting_condition,
            angle_of_impact=angle,
            environment=environment,
            driver_profile=driver_profile,
            interventions=intervention_plan.interventions if apply_interventions and intervention_plan.interventions else None,
            verbose=False
        )

        results.append({
            "severity": severity,
            "risk_score": risk_score,
            "location": location,
            "impact_force": impact_force
        })

    if not results:
        return {}

    df = pd.DataFrame(results)
    distribution = df['severity'].value_counts().to_dict()
    distribution['average_risk_score'] = float(round(df['risk_score'].mean(), 2)) if not df['risk_score'].isna().all() else None
    distribution['average_impact_force'] = float(round(df['impact_force'].mean(), 2))
    return distribution


def run_monte_carlo_intervention_analysis(runs: int = 20) -> dict:
    """Compare baseline vs intervention outcomes using identical random scenarios."""

    baseline_records: list[dict] = []
    intervention_records: list[dict] = []

    for _ in range(runs):
        cause, location, primary_vehicle_type, road_condition, lighting_condition, weather_condition, human_factor = sample_from_data_sources()
        environment = build_environment_state(location, road_condition, lighting_condition, weather_condition)
        driver_profile = DriverProfile.from_factor(human_factor)
        intervention_plan = suggest_interventions(location, cause, environment.road_condition, environment.lighting_condition, human_factor)
        accident_type = random.choice(list(accident_types.keys()))
        vehicle1_spec, vehicle2_spec, pedestrian_spec, angle = prepare_accident_entities(accident_type, primary_vehicle_type)

        # Baseline run
        vehicle1_base, vehicle2_base, pedestrian_base = instantiate_entities(vehicle1_spec, vehicle2_spec, pedestrian_spec)
        _, _, _, _, _, force_base, severity_base, risk_base, _ = simulate_collision(
            vehicle1_base,
            vehicle2_base,
            pedestrian=pedestrian_base,
            accident_type=accident_type,
            road_condition=environment.road_condition,
            lighting_condition=environment.lighting_condition,
            angle_of_impact=angle,
            environment=environment,
            driver_profile=driver_profile,
            interventions=None,
            verbose=False
        )
        baseline_records.append({
            "severity": severity_base,
            "risk_score": risk_base,
            "impact_force": force_base
        })

        # Intervention run (if plan active)
        vehicle1_int, vehicle2_int, pedestrian_int = instantiate_entities(vehicle1_spec, vehicle2_spec, pedestrian_spec)
        _, _, _, _, _, force_int, severity_int, risk_int, _ = simulate_collision(
            vehicle1_int,
            vehicle2_int,
            pedestrian=pedestrian_int,
            accident_type=accident_type,
            road_condition=environment.road_condition,
            lighting_condition=environment.lighting_condition,
            angle_of_impact=angle,
            environment=environment,
            driver_profile=driver_profile,
            interventions=intervention_plan.interventions if intervention_plan.interventions else None,
            verbose=False
        )
        intervention_records.append({
            "severity": severity_int,
            "risk_score": risk_int,
            "impact_force": force_int
        })

    def summarize(records: list[dict]) -> dict:
        df = pd.DataFrame(records)
        distribution = df['severity'].value_counts().to_dict()
        distribution['average_risk_score'] = float(round(df['risk_score'].mean(), 2)) if not df['risk_score'].isna().all() else None
        distribution['average_impact_force'] = float(round(df['impact_force'].mean(), 2))
        return distribution

    if not baseline_records:
        return {}

    return {
        'baseline': summarize(baseline_records),
        'intervention': summarize(intervention_records)
    }


def generate_recommendations(cause, road_condition, lighting_condition, accident_type):
    """Generate randomized recommendations based on accident factors"""
    recs = []
    
    # Add cause-specific recommendations
    if cause in recommendations_db:
        recs.extend(random.sample(recommendations_db[cause], min(2, len(recommendations_db[cause]))))
    
    # Add road condition recommendations
    if road_condition in recommendations_db:
        recs.extend(random.sample(recommendations_db[road_condition], min(1, len(recommendations_db[road_condition]))))
    
    # Add lighting recommendations
    if lighting_condition in recommendations_db:
        recs.extend(random.sample(recommendations_db[lighting_condition], min(1, len(recommendations_db[lighting_condition]))))
    
    # Add accident type recommendations
    if accident_type in recommendations_db:
        recs.extend(random.sample(recommendations_db[accident_type], min(1, len(recommendations_db[accident_type]))))
    
    # Remove duplicates and limit to 4-5 recommendations
    recs = list(set(recs))[:5]
    
    return recs

# Load historical accident data from CSV (based on PDF Materials and Methods A. Data Sources)
# Data sources: local government offices, barangay authorities, police reports, LTO
# Includes: causes, time/location, vehicle types, weather/lighting, road characteristics, human factors
try:
    accident_data = pd.read_csv('accident_data.csv')
except FileNotFoundError:
    # Create default accident data if file doesn't exist
    default_data = [
        ['Overspeeding', '2023-01-15 14:30', 'Bayabas', 'Car', 'Sunny', 'Daylight', 'Dry road', 'Fatigue', 'Moderate'],
        ['Human Error', '2023-02-20 22:00', 'Main Highway', 'Motorcycle', 'Rainy', 'Night', 'Slippery surface', 'Drunk Driving', 'Severe'],
        ['Mechanical Failure', '2023-03-10 08:15', 'Barangay 1', 'Truck', 'Cloudy', 'Daylight', 'Blind curve', 'None', 'Minor'],
        ['Fatigue', '2023-04-05 23:45', 'Bayabas', 'Car', 'Clear', 'Night', 'Dark area', 'Sleepy driver', 'Moderate'],
        ['Drunk Driving', '2023-05-12 01:20', 'Main Highway', 'Car', 'Rainy', 'Night', 'Wet road', 'Under influence', 'Severe']
    ]
    accident_data = pd.DataFrame(default_data, columns=['cause', 'time', 'location', 'vehicle_type', 'weather', 'lighting', 'road_characteristics', 'human_factors', 'severity'])
    accident_data.to_csv('accident_data.csv', index=False)
    print("Created default accident_data.csv with sample data")

def sample_from_data_sources():
    """Sample accident parameters from historical data sources"""
    global accident_data
    try:
        accident_data = pd.read_csv('accident_data.csv')
    except Exception:
        pass
    row = accident_data.sample().iloc[0]

    # Map data to simulation parameters
    cause = row.get('cause', 'Overspeeding')
    location = row.get('location', 'Bayabas')

    # Map vehicle type
    vehicle_type_map = {
        'Car': 'Car',
        'Motorcycle': 'Motorcycle',
        'Truck': 'Truck',
        'Bus': 'Bus'
    }
    vehicle_type = vehicle_type_map.get(row.get('vehicle_type', 'Car'), 'Car')

    # Map road condition
    road_char = str(row.get('road_characteristics', 'Dry road')).lower()
    if 'slippery' in road_char:
        road_condition = 'slippery'
    elif 'wet' in road_char:
        road_condition = 'wet'
    else:
        road_condition = 'dry'

    # Map lighting condition
    lighting_map = {
        'Night': 'poor',
        'Dusk': 'poor',
        'Daylight': 'good'
    }
    lighting_condition = lighting_map.get(row.get('lighting', 'Daylight'), 'good')

    # Weather information
    weather_raw = str(row.get('weather', 'Sunny'))
    weather_condition = weather_raw.title()
    if 'rainy' in weather_raw.lower() and road_condition == 'dry':
        road_condition = 'wet'

    human_factor = str(row.get('human_factors', 'None'))

    return cause, location, vehicle_type, road_condition, lighting_condition, weather_condition, human_factor

class Vehicle:
    def __init__(self, mass, initial_velocity, position=0, name="Vehicle"):
        self.mass = mass  # kg
        self.velocity = initial_velocity  # m/s
        self.position = position  # m
        self.name = name
        self.acceleration = 0  # m/s^2
        self.braking = False
        self.reaction_time = 0  # seconds
        self.braking_timer = 0

    def update_position(self, dt):
        if self.braking and self.braking_timer <= 0:
            self.acceleration = -self.friction_coeff * 9.81  # braking deceleration
        self.position += self.velocity * dt + 0.5 * self.acceleration * dt**2
        self.velocity += self.acceleration * dt
        # Only stop at zero when braking (not for initial negative velocities in head-on collisions)
        if self.braking and self.velocity < 0:
            self.velocity = 0  # stop at zero only when braking
        if self.braking_timer > 0:
            self.braking_timer -= dt

    def apply_braking(self, friction_coeff, reaction_time=0):
        self.friction_coeff = friction_coeff
        self.reaction_time = reaction_time
        self.braking_timer = reaction_time
        self.braking = True

class Pedestrian:
    def __init__(self, mass=70, position=0, name="Pedestrian"):
        self.mass = mass  # kg
        self.position = position  # m
        self.name = name
        self.velocity = 0  # pedestrians don't move in this simple model

    def update_position(self, dt):
        # Pedestrians don't move
        pass

def simulate_collision(
    vehicle1,
    vehicle2=None,
    pedestrian=None,
    accident_type="rear-end",
    dt=0.01,
    total_time=10.0,
    road_condition="dry",
    lighting_condition="good",
    angle_of_impact=0,
    environment: EnvironmentState | None = None,
    driver_profile: DriverProfile | None = None,
    interventions: list[str] | None = None,
    verbose: bool = True
):
    """
    Simulate the motion of two vehicles until collision or time out.
    Assume vehicle1 is behind vehicle2 for rear-end collision.
    angle_of_impact: 0 for rear-end, 180 for head-on, etc. (in degrees)
    """
    # Set friction and reaction using environment-aware modeling
    if environment:
        friction_coeff = environment.get_effective_friction(driver_profile)
        reaction_time = environment.get_effective_reaction(driver_profile)
    else:
        friction_map = {"dry": 0.8, "wet": 0.5, "slippery": 0.1}
        friction_coeff = friction_map.get(road_condition, 0.8)
        reaction_map = {"good": 0.5, "poor": 1.5}
        reaction_time = reaction_map.get(lighting_condition, 0.5)

    # Apply interventions effects before simulation starts
    if interventions:
        for intervention in interventions:
            effects = intervention_effects.get(intervention)
            if not effects:
                continue
            if 'lighting_override' in effects and environment:
                environment = EnvironmentState(
                    road_condition=environment.road_condition,
                    lighting_condition=effects['lighting_override'],
                    weather_condition=environment.weather_condition,
                    slope=environment.slope,
                    curvature=environment.curvature
                )
                reaction_time = environment.get_effective_reaction(driver_profile)
            if 'road_condition_override' in effects and environment:
                environment = EnvironmentState(
                    road_condition=effects['road_condition_override'],
                    lighting_condition=environment.lighting_condition,
                    weather_condition=environment.weather_condition,
                    slope=environment.slope,
                    curvature=environment.curvature
                )
                friction_coeff = environment.get_effective_friction(driver_profile)
            if 'friction_bonus' in effects:
                friction_coeff = min(1.2, friction_coeff + effects['friction_bonus'])
            if 'reaction_multiplier' in effects:
                reaction_time *= effects['reaction_multiplier']
            if 'driver_risk_multiplier' in effects and driver_profile:
                driver_profile = DriverProfile(
                    factor=driver_profile.factor,
                    reaction_multiplier=driver_profile.reaction_multiplier,
                    braking_multiplier=driver_profile.braking_multiplier,
                    risk_multiplier=driver_profile.risk_multiplier * effects['driver_risk_multiplier'],
                    notes=driver_profile.notes
                )
            if 'speed_reduction' in effects:
                reduction = max(0.0, min(0.6, effects['speed_reduction']))
                if vehicle1:
                    vehicle1.velocity *= (1 - reduction)
                if vehicle2:
                    vehicle2.velocity *= (1 - reduction * 0.5)

    reaction_time = max(0.2, min(reaction_time, 3.5))
    friction_coeff = max(0.05, min(friction_coeff, 1.2))

    # Apply braking only for rear-end collisions
    if accident_type == "rear-end" and vehicle1:
        vehicle1.apply_braking(friction_coeff, reaction_time)
    
    # Set up entities based on accident type
    entities = [vehicle1]
    if vehicle2:
        entities.append(vehicle2)
    if pedestrian:
        entities.append(pedestrian)
    
    # Initialize positions and velocities lists
    positions = {entity.name: [entity.position] for entity in entities}
    velocities = {entity.name: [entity.velocity] for entity in entities}
    
    time = 0
    collision_distance = 5  # meters
    impact_force = 0
    severity = "none"
    
    while time < total_time:
        # Update all entities
        for entity in entities:
            if hasattr(entity, 'update_position'):
                entity.update_position(dt)
            # Pedestrian doesn't move
        
        # Record positions and velocities
        for entity in entities:
            positions[entity.name].append(entity.position)
            velocities[entity.name].append(entity.velocity)
        
        # Check for collision based on type
        collided = False
        if accident_type == "rear-end" and vehicle2:
            if vehicle1.position >= vehicle2.position - collision_distance and vehicle1.velocity > vehicle2.velocity:
                collided = True
        elif accident_type == "head-on" and vehicle2:
            if abs(vehicle1.position - vehicle2.position) < collision_distance and vehicle1.velocity > 0 and vehicle2.velocity < 0:
                collided = True
        elif accident_type == "side-impact" and vehicle2:
            if vehicle1.position >= vehicle2.position - collision_distance:
                collided = True
        elif accident_type == "pedestrian" and pedestrian:
            if vehicle1.position >= pedestrian.position - collision_distance:
                collided = True
        
        if collided:
            if verbose:
                print(f"Collision detected at time {time:.2f}s, position {vehicle1.position:.2f}m")
            # Calculate impact force
            if accident_type == "pedestrian":
                v1_initial = velocities[vehicle1.name][-1]  # current velocity before collision
                total_mass = vehicle1.mass + pedestrian.mass
                v_final = (vehicle1.mass * v1_initial + pedestrian.mass * pedestrian.velocity) / total_mass
                delta_v = abs(v1_initial - v_final)
                impact_force = vehicle1.mass * delta_v / dt
                vehicle1.velocity = v_final
                pedestrian.velocity = v_final
            else:
                # Vehicle-vehicle collision
                v1_before = velocities[vehicle1.name][-2] if len(velocities[vehicle1.name]) > 1 else velocities[vehicle1.name][-1]
                total_mass = vehicle1.mass + vehicle2.mass
                v_final = (vehicle1.mass * vehicle1.velocity + vehicle2.mass * vehicle2.velocity) / total_mass
                vehicle1.velocity = v_final
                vehicle2.velocity = v_final
                delta_v = abs(vehicle1.velocity - v1_before)
                impact_force = vehicle1.mass * delta_v / dt
            
            # Adjust impact force based on environmental modifiers and angle of impact
            slope_factor = 1.0
            curvature_factor = 1.0
            if environment:
                slope_factor += abs(environment.slope) / 50.0
                if environment.curvature in {"blind_curve", "sharp_turn"}:
                    curvature_factor += 0.2
            angle_factor = 1 + abs(np.sin(np.deg2rad(angle_of_impact))) * 0.3
            impact_force *= slope_factor * curvature_factor * angle_factor

            # Classify severity
            if impact_force < 50000:
                severity = "minor"
            elif impact_force < 150000:
                severity = "moderate"
            else:
                severity = "severe"
            if verbose:
                print(f"Impact force: {impact_force:.2f} N, Severity: {severity}")
            break
        
        time += dt
    
    # Create DataFrame
    max_len = max(len(pos) for pos in positions.values())
    time_array = np.arange(0, max_len) * dt
    sim_data = pd.DataFrame({'time': time_array})
    for name in positions:
        sim_data[f'position_{name.lower().replace(" ", "_")}'] = positions[name] + [positions[name][-1]] * (max_len - len(positions[name]))
        sim_data[f'velocity_{name.lower().replace(" ", "_")}'] = velocities[name] + [velocities[name][-1]] * (max_len - len(velocities[name]))
    
    # For compatibility, set pos1, pos2, vel1, vel2 to vehicle1 and vehicle2 or pedestrian
    pos1 = positions.get(vehicle1.name, [])
    vel1 = velocities.get(vehicle1.name, [])
    if vehicle2:
        pos2 = positions.get(vehicle2.name, [])
        vel2 = velocities.get(vehicle2.name, [])
    elif pedestrian:
        pos2 = positions.get(pedestrian.name, [])
        vel2 = velocities.get(pedestrian.name, [])
    else:
        pos2 = []
        vel2 = []
    
    risk_score = calculate_risk_score(impact_force, severity, environment, driver_profile)

    return np.array(pos1), np.array(pos2), np.array(vel1), np.array(vel2), time, impact_force, severity, risk_score, sim_data

def print_simulation_steps(accident_type, cause, location, road_condition, lighting_condition, vehicle1, vehicle2=None, pedestrian=None, weather=None, driver_profile: DriverProfile | None = None, environment: EnvironmentState | None = None, interventions: InterventionPlan | None = None):
    """Print step-by-step breakdown of the road accident simulation process"""
    print("\n" + "="*80)
    print("ROAD ACCIDENT SIMULATION STEPS")
    print("="*80)
    
    print(f"\nStep 1: Accident Scenario Setup")
    print(f"   - Accident Type: {accident_types[accident_type]['description']}")
    print(f"   - Location: {location}, Labo, Camarines Norte")
    print(f"   - Primary Cause: {cause}")
    print(f"   - Road Conditions: {road_condition}")
    print(f"   - Lighting Conditions: {lighting_condition}")
    if weather:
        print(f"   - Weather: {weather}")
    if environment:
        print(f"   - Road Geometry: {environment.curvature}, Approx. Slope: {environment.slope:.1f}°")
    if driver_profile:
        print(f"   - Driver State: {driver_profile.factor} ({driver_profile.notes})")
    
    print(f"\nStep 2: Vehicle/Pedestrian Initialization")
    if vehicle1:
        print(f"   - Vehicle 1: {vehicle1.name} (Mass: {vehicle1.mass} kg, Initial Speed: {vehicle1.velocity*3.6:.1f} km/h)")
    if vehicle2:
        print(f"   - Vehicle 2: {vehicle2.name} (Mass: {vehicle2.mass} kg, Initial Speed: {vehicle2.velocity*3.6:.1f} km/h)")
    if pedestrian:
        print(f"   - Pedestrian: {pedestrian.name} (Mass: {pedestrian.mass} kg, Position: {pedestrian.position:.1f} m)")
    
    print(f"\nStep 3: Environmental Factors Applied")
    friction_map = {"dry": 0.8, "wet": 0.5, "slippery": 0.1}
    reaction_map = {"good": 0.5, "poor": 1.5}
    friction_coeff = friction_map.get(road_condition, 0.8)
    reaction_time = reaction_map.get(lighting_condition, 0.5)
    if environment:
        friction_coeff = environment.get_effective_friction(driver_profile)
        reaction_time = environment.get_effective_reaction(driver_profile)
    print(f"   - Friction Coefficient: {friction_coeff} (based on {road_condition} road)")
    print(f"   - Driver Reaction Time: {reaction_time}s (based on {lighting_condition} lighting)")
    
    if accident_type == "rear-end":
        print(f"   - Braking Applied: Vehicle 1 applies brakes after {reaction_time}s reaction time")
    
    print(f"\nStep 4: Physics Simulation Begins")
    print(f"   - Time Step: 0.01 seconds")
    print(f"   - Total Simulation Time: 10.0 seconds")
    print(f"   - Collision Detection Distance: 5.0 meters")
    
    print(f"\nStep 5: Motion Calculation")
    print(f"   - Position updates using kinematic equations:")
    print(f"     position = position + velocity * time + 0.5 * acceleration * time^2")
    print(f"     velocity = velocity + acceleration * time")
    
    if accident_type == "rear-end":
        deceleration = friction_coeff * 9.81
        print(f"   - Braking Deceleration: {deceleration:.1f} m/s² (friction * gravity)")
    
    print(f"\nStep 6: Collision Detection")
    if accident_type == "rear-end":
        print(f"   - Checking if Vehicle 1 overtakes Vehicle 2 within collision distance")
    elif accident_type == "head-on":
        print(f"   - Checking if vehicles meet within collision distance")
    elif accident_type == "side-impact":
        print(f"   - Checking perpendicular collision proximity")
    elif accident_type == "pedestrian":
        print(f"   - Checking if vehicle reaches pedestrian position")
    
    print(f"\nStep 7: Impact Analysis (if collision occurs)")
    print(f"   - Momentum Conservation: m1*v1 + m2*v2 = (m1+m2)*v_final")
    print(f"   - Impact Force Calculation: F = m * delta_v / delta_t")
    print(f"   - Severity Classification:")
    print(f"     • Minor: Impact Force < 50,000 N")
    print(f"     • Moderate: 50,000 N <= Impact Force < 150,000 N")
    print(f"     • Severe: Impact Force >= 150,000 N")
    
    print(f"\nStep 8: Data Recording")
    print(f"   - Position and velocity data collected every 0.01 seconds")
    print(f"   - Simulation data stored in pandas DataFrame")
    
    print(f"\nStep 9: Report Generation")
    print(f"   - Accident analysis and recommendations generated")
    print(f"   - Excel report created with formatted data")
    
    print(f"\nStep 10: Visualization")
    print(f"   - Position vs time plots generated for analysis")
    if interventions and interventions.interventions:
        print(f"\nStep 11: Intervention Modeling")
        for item in interventions.interventions:
            desc = intervention_effects.get(item, {}).get('description', 'Applied based on field recommendations')
            print(f"   - {item.replace('_', ' ').title()}: {desc}")
    
    print("\n" + "="*80)
    print("SIMULATION EXECUTION BEGINS")
    print("="*80 + "\n")

def generate_report(
    vehicle1,
    vehicle2,
    pedestrian,
    accident_type,
    collision_time,
    impact_force,
    severity,
    road_condition,
    lighting,
    initial_v1,
    initial_v2,
    cause,
    location,
    angle_of_impact=0,
    weather=None,
    driver_profile: DriverProfile | None = None,
    environment: EnvironmentState | None = None,
    risk_score: float | None = None,
    intervention_plan: InterventionPlan | None = None,
    validation_text: str | None = None,
    comparison: dict | None = None
):
    accident_desc = accident_types[accident_type]["description"]
    geometry_summary = (
        f"Road Geometry: {environment.curvature} (Slope: {environment.slope:.1f}°)" if environment else "Road Geometry: Not specified"
    )
    report = f"""
Road Accident Simulation Report
================================

Accident Type: {accident_desc}
Location: {location}, Labo, Camarines Norte
Time of Accident: {collision_time:.2f} seconds into simulation
Cause: {cause}
Weather: {weather or 'Not recorded'}
{geometry_summary}

Vehicles Involved:
"""
    if vehicle1:
        report += f"- {vehicle1.name}: Mass = {vehicle1.mass} kg, Initial Speed = {initial_v1 * 3.6:.1f} km/h\n"
    if vehicle2:
        report += f"- {vehicle2.name}: Mass = {vehicle2.mass} kg, Initial Speed = {initial_v2 * 3.6:.1f} km/h\n"
    if pedestrian:
        report += f"- {pedestrian.name}: Mass = {pedestrian.mass} kg\n"
    
    recommendations = generate_recommendations(cause, road_condition, lighting, accident_type)
    recommendations_text = "\n".join(f"- {rec}" for rec in recommendations)
    system_summary = generate_systemic_summary(cause, driver_profile, environment, intervention_plan)
    validation_line = validation_text or "Historical validation unavailable."

    baseline_severity = comparison['baseline']['severity'] if comparison else severity
    intervention_severity = comparison['intervention']['severity'] if comparison else severity
    baseline_force = comparison['baseline']['impact_force'] if comparison else impact_force
    intervention_force = comparison['intervention']['impact_force'] if comparison else impact_force
    baseline_risk = comparison['baseline'].get('risk_score') if comparison else risk_score
    intervention_risk = comparison['intervention'].get('risk_score') if comparison else risk_score
    impact_force_reduction = None
    if comparison and baseline_force:
        impact_force_reduction = (baseline_force - intervention_force) / baseline_force * 100
    risk_reduction = None
    if comparison and baseline_risk:
        risk_reduction = baseline_risk - intervention_risk
    baseline_risk_formatted = f"{baseline_risk:.2f}" if baseline_risk is not None else "N/A"
    intervention_risk_formatted = f"{intervention_risk:.2f}" if intervention_risk is not None else "N/A"
    risk_change_text = (
        f"{risk_reduction:.2f} (Baseline {baseline_risk_formatted}, Intervention {intervention_risk_formatted})"
        if risk_reduction is not None and baseline_risk is not None and intervention_risk is not None
        else f"0.00 (Baseline {baseline_risk_formatted}, Intervention {intervention_risk_formatted})"
    )
    force_reduction_text = f"{impact_force_reduction:.2f}%" if impact_force_reduction is not None else "0%"
    intervention_desc = intervention_plan.describe() if intervention_plan else 'None'
    baseline_risk_text = baseline_risk_formatted
    intervention_risk_text = intervention_risk_formatted
    
    report += f"""

Analysis:
This simulation demonstrates how speed, road conditions, and lighting contribute to accidents.
Severity is classified based on impact force: minor (<50kN), moderate (50-150kN), severe (>150kN).
Recommendations:
{recommendations_text}

Systems Interaction Summary:
- {system_summary}

Validation Against Historical Records:
- {validation_line}

Intervention Scenario Comparison:
- Baseline Severity: {baseline_severity}
- Intervention Severity: {intervention_severity}
- Baseline Impact Force: {baseline_force:.2f} N
- Intervention Impact Force: {intervention_force:.2f} N
- Risk Score Change: {risk_change_text}
- Impact Force Reduction: {force_reduction_text}
- Interventions Applied: {intervention_desc}

"""
    # Create formatted Excel report
    wb = Workbook()
    ws = wb.active
    ws.title = "Accident Report"
    
    # Add title
    ws['A1'] = "Road Accident Simulation Report"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:O1')
    
    # Add headers starting from row 3
    data = {
        'Accident Type': accident_desc,
        'Location': f"{location}, Labo, Camarines Norte",
        'Time of Accident': f"{collision_time:.2f} seconds",
        'Cause': cause,
        'Vehicle 1 Name': vehicle1.name if vehicle1 else '',
        'Vehicle 1 Mass (kg)': vehicle1.mass if vehicle1 else '',
        'Vehicle 1 Initial Speed (km/h)': f"{initial_v1 * 3.6:.1f}" if initial_v1 else '',
        'Vehicle 2 Name': vehicle2.name if vehicle2 else '',
        'Vehicle 2 Mass (kg)': vehicle2.mass if vehicle2 else '',
        'Vehicle 2 Initial Speed (km/h)': f"{initial_v2 * 3.6:.1f}" if initial_v2 else '',
        'Pedestrian Name': pedestrian.name if pedestrian else '',
        'Pedestrian Mass (kg)': pedestrian.mass if pedestrian else '',
        'Road Conditions': road_condition,
        'Lighting Conditions': lighting,
        'Weather': weather or 'Not recorded',
        'Angle of Impact (degrees)': angle_of_impact,
        'Collision Time (s)': f"{collision_time:.2f}",
        'Impact Force (N)': f"{impact_force:.2f}",
        'Severity': severity,
        'Risk Score (0-10)': risk_score if risk_score is not None else baseline_risk,
        'Analysis': 'This simulation demonstrates how speed, road conditions, and lighting contribute to accidents.',
        'Recommendations': recommendations_text,
        'Systems Summary': system_summary,
        'Historical Validation': validation_line,
        'Interventions Applied': intervention_desc,
        'Baseline Severity': baseline_severity,
        'Intervention Severity': intervention_severity,
        'Baseline Impact Force (N)': f"{baseline_force:.2f}",
        'Intervention Impact Force (N)': f"{intervention_force:.2f}",
        'Impact Force Reduction (%)': force_reduction_text,
        'Risk Score Change': risk_change_text,
        'Baseline Risk Score': baseline_risk_text,
        'Intervention Risk Score': intervention_risk_text
    }
    
    headers = list(data.keys())
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
        ws.cell(row=3, column=col).font = Font(bold=True)
        ws.cell(row=3, column=col).fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold color
        ws.cell(row=3, column=col).alignment = Alignment(horizontal='center')
    
    # Add data
    for col, value in enumerate(data.values(), 1):
        ws.cell(row=4, column=col, value=value)
        ws.cell(row=4, column=col).alignment = Alignment(horizontal='center')
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for row in range(1, 5):  # Check rows 1 to 4
            cell = ws.cell(row=row, column=col_num)
            if cell.value and not isinstance(cell, MergedCell):
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 30)  # Cap at 30 for readability
        ws.column_dimensions[column_letter].width = adjusted_width
    
    try:
        wb.save('accident_report.xlsx')
        print("Formatted Excel report saved as 'accident_report.xlsx'")
    except PermissionError:
        print("Could not save Excel report due to permission error. Report printed above.")
    
    return report


def create_labo_roads_geojson(force_regenerate: bool = False) -> Path | None:
    """Ensure a clipped Labo road network GeoJSON exists and return its path."""

    base_dir = Path(__file__).resolve().parent if '__file__' in globals() else Path(os.getcwd())
    gis_dir = base_dir / "GIS"
    output_geojson_path = gis_dir / "labo_roads.geojson"

    if output_geojson_path.exists() and not force_regenerate:
        print(f"Found existing {output_geojson_path}. Reusing it for the map overlay.")
        return output_geojson_path

    gis_dir.mkdir(parents=True, exist_ok=True)

    def pick_source(candidates: list[Path]) -> Path | None:
        for candidate in candidates:
            if candidate.exists():
                return candidate
        return None

    road_source = pick_source([
        gis_dir / "highway.shp",
        gis_dir / "Labo_Roads.shp",
        gis_dir / "labo_roads.geojson.shp"
    ])

    boundary_source = pick_source([
        gis_dir / "Labo_Boundary.shp",
        gis_dir / "Labo_Boundary.geojson",
        gis_dir / "Labo_Boundary.shp.gpkg"
    ])

    if road_source is None:
        print("No road network file found in the GIS directory. Expected 'highway.shp' or a Labo-specific layer.")
        return None

    print(f"Reading road network from {road_source}...")
    try:
        roads = gpd.read_file(road_source)
    except Exception as exc:
        print(f"Error reading road network layer: {exc}")
        return None

    clip_required = boundary_source is not None and road_source.name.lower() not in {"labo_roads.geojson.shp", "labo_roads.shp"}

    if clip_required:
        print(f"Reading Labo boundary from {boundary_source}...")
        try:
            boundary = gpd.read_file(boundary_source)
        except Exception as exc:
            print(f"Error reading boundary layer: {exc}")
            return None

        print("Clipping roads to the Labo boundary...")
        try:
            roads = gpd.clip(roads, boundary)
        except Exception as exc:
            print(f"Error during clipping: {exc}")
            return None
    else:
        if boundary_source is None:
            print("Boundary layer not found; exporting the available road layer as-is.")
        else:
            print("Detected an existing Labo-only road layer; skipping additional clipping.")

    if roads.empty:
        print("No road geometries found to export. GeoJSON will not be created.")
        return None

    print(f"Saving Labo road network to {output_geojson_path}...")
    try:
        roads.to_file(output_geojson_path, driver="GeoJSON")
    except Exception as exc:
        print(f"Error saving GeoJSON: {exc}")
        return None

    print("Success! labo_roads.geojson is ready.")
    return output_geojson_path


def create_map_visualization(
    sim_data,
    vehicle1_name,
    vehicle2_name=None,
    pedestrian_name=None,
    location="Labo, Camarines Norte",
    labo_geojson_path: Path | None = None,
    collision_time: float | None = None
):
    """
    Creates an animated map visualization of the road accident simulation.
    """
    print("\nStarting map visualization process...")
    
    # 1. Prioritize GIS road network for accurate alignment
    candidate_geojson = Path(labo_geojson_path) if labo_geojson_path else (
        (Path(__file__).resolve().parent if '__file__' in globals() else Path(os.getcwd())) / "GIS" / "labo_roads.geojson"
    )

    roads_gdf_latlon: gpd.GeoDataFrame | None = None
    roads_geojson_data: dict | None = None
    if candidate_geojson.exists():
        try:
            roads_gdf_latlon = gpd.read_file(candidate_geojson)
            if roads_gdf_latlon.empty:
                print(f"Labo roads file {candidate_geojson} contains no features; falling back to OSM data.")
                roads_gdf_latlon = None
            else:
                if roads_gdf_latlon.crs is None:
                    roads_gdf_latlon = roads_gdf_latlon.set_crs('EPSG:4326', allow_override=True)
                else:
                    roads_gdf_latlon = roads_gdf_latlon.to_crs('EPSG:4326')
                roads_geojson_data = json.loads(roads_gdf_latlon.to_json())
        except Exception as exc:
            print(f"Could not read Labo roads GeoJSON at {candidate_geojson}: {exc}")
            roads_gdf_latlon = None
            roads_geojson_data = None
    else:
        print(f"Labo roads GeoJSON not found at {candidate_geojson}")

    map_center = [14.156, 122.83]  # Default to Labo town center coordinates
    streets = None
    if roads_gdf_latlon is not None:
        try:
            roads_centroid_merc = roads_gdf_latlon.to_crs('EPSG:3857').geometry.unary_union.centroid
            centroid_latlon = gpd.GeoSeries([roads_centroid_merc], crs='EPSG:3857').to_crs('EPSG:4326').iloc[0]
            map_center = [centroid_latlon.y, centroid_latlon.x]
            print(f"Map centered using GIS road network ({len(roads_gdf_latlon)} features)")
        except Exception as exc:
            print(f"Could not derive centroid from GIS roads: {exc}")
    else:
        # Fallback to OSM data for broader context
        place_name = "Camarines Norte, Philippines"
        ox.settings.timeout = 30  # Set timeout to 30 seconds
        try:
            graph = ox.graph_from_place(place_name, network_type='drive')
            nodes, streets = ox.graph_to_gdfs(graph)

            labo_area = nodes[(nodes.geometry.y >= 14.1) & (nodes.geometry.y <= 14.25) &
                             (nodes.geometry.x >= 122.7) & (nodes.geometry.x <= 122.95)]

            if not labo_area.empty:
                labo_center = labo_area.to_crs('EPSG:3857').geometry.unary_union.centroid
                centroid_latlon = gpd.GeoSeries([labo_center], crs='EPSG:3857').to_crs('EPSG:4326').iloc[0]
                map_center = [centroid_latlon.y, centroid_latlon.x]
                print(f"Successfully fetched map data for Camarines Norte, focusing on Labo area ({len(labo_area)} nodes)")
            else:
                broader_center = nodes.to_crs('EPSG:3857').geometry.unary_union.centroid
                centroid_latlon = gpd.GeoSeries([broader_center], crs='EPSG:3857').to_crs('EPSG:4326').iloc[0]
                map_center = [centroid_latlon.y, centroid_latlon.x]
                print(f"Using broader Camarines Norte area for Labo roads ({len(nodes)} nodes)")
        except Exception as e:
            print(f"Could not fetch map data for {place_name}. Using default Labo coordinates. Error: {e}")
            streets = None

    # Create a base map
    m = folium.Map(location=map_center, zoom_start=15)

    if roads_geojson_data is not None:
        def road_style(_):
            return {"color": "#ff7f0e", "weight": 3, "opacity": 0.85}

        folium.GeoJson(
            roads_geojson_data,
            name="Labo Road Network",
            style_function=road_style
        ).add_to(m)
        folium.LayerControl().add_to(m)
        print(f"Added Labo roads overlay from {candidate_geojson}")

        try:
            roads_merc_for_3d = roads_gdf_latlon.to_crs('EPSG:3857') if roads_gdf_latlon is not None else None
            deck_data: list[dict] = []
            if roads_gdf_latlon is not None and roads_merc_for_3d is not None and not roads_gdf_latlon.empty:
                lengths = roads_merc_for_3d.geometry.length
                max_length = float(lengths.max()) if len(lengths) else 1.0

                def register_path(line_latlon: LineString, line_merc: LineString) -> None:
                    coords = [[float(pt[0]), float(pt[1])] for pt in line_latlon.coords]
                    if len(coords) < 2:
                        return
                    relative = float(line_merc.length) / max_length if max_length else 0.1
                    elevation = max(15.0, relative * 180.0)
                    deck_data.append({"coordinates": coords, "elevation": elevation})

                for geom_latlon, geom_merc in zip(roads_gdf_latlon.geometry, roads_merc_for_3d.geometry):
                    if isinstance(geom_latlon, LineString):
                        register_path(geom_latlon, geom_merc)
                    elif isinstance(geom_latlon, MultiLineString):
                        for segment_latlon, segment_merc in zip(geom_latlon.geoms, geom_merc.geoms if hasattr(geom_merc, 'geoms') else []):
                            if isinstance(segment_latlon, LineString):
                                register_path(segment_latlon, segment_merc)

            if deck_data:
                deck_data_json = json.dumps(deck_data)
                deck_script = f"""
                <script>
                (function() {{
                    const map = {m.get_name()};
                    const deckData = {deck_data_json};
                    const deckScriptUrl = 'https://unpkg.com/deck.gl@8.9.36/dist.min.js';
                    const leafletDeckUrl = 'https://unpkg.com/@deck.gl/leaflet@8.9.36/dist.min.js';

                    function ensureScript(url) {{
                        return new Promise(function(resolve, reject) {{
                            if (document.querySelector('script[src="' + url + '"]')) {{
                                resolve();
                                return;
                            }}
                            const script = document.createElement('script');
                            script.src = url;
                            script.async = true;
                            script.onload = resolve;
                            script.onerror = reject;
                            document.head.appendChild(script);
                        }});
                    }}

                    Promise.all([ensureScript(deckScriptUrl), ensureScript(leafletDeckUrl)])
                        .then(function() {{
                            if (!window.deck || !deck.LeafletLayer) {{
                                console.error('deck.gl Leaflet integration unavailable.');
                                return;
                            }}

                            const roadLayer = new deck.PathLayer({{
                                id: 'labo-roads-3d',
                                data: deckData,
                                getPath: d => d.coordinates,
                                getColor: d => [31, 119, 180, 210],
                                widthScale: 20,
                                widthMinPixels: 3,
                                getWidth: 4,
                                extruded: true,
                                getElevation: d => d.elevation,
                                elevationScale: 1,
                                material: true,
                                pickable: false
                            }});

                            if (map.__deckRoadLayer) {{
                                map.removeLayer(map.__deckRoadLayer);
                            }}

                            const deckLayer = new deck.LeafletLayer({{
                                layers: [roadLayer],
                                views: [new deck.MapView({{ repeat: true }})]
                            }});

                            deckLayer.addTo(map);
                            map.__deckRoadLayer = deckLayer;
                        }})
                        .catch(function(err) {{
                            console.error('Failed to initialise deck.gl roads overlay.', err);
                        }});
                }})();
                </script>
                """
                m.get_root().html.add_child(folium.Element(deck_script))
                print(f"Added deck.gl 3D path overlay with {len(deck_data)} segments")
        except Exception as exc:
            print(f"Unable to build 3D road overlay: {exc}")

    # Add accident location marker with animated pulse to draw attention
    accident_marker_css = """
    <style>
    .accident-marker {
        position: relative;
        width: 34px;
        height: 34px;
    }
    .accident-marker-core {
        position: absolute;
        top: 50%;
        left: 50%;
        width: 18px;
        height: 18px;
        transform: translate(-50%, -50%);
        background: #d9534f;
        border-radius: 50%;
        color: #ffffff;
        display: flex;
        align-items: center;
        justify-content: center;
        box-shadow: 0 0 10px rgba(217, 83, 79, 0.7);
        animation: accident-core-pulse 1.6s infinite;
    }
    .accident-marker-core i {
        font-size: 12px;
    }
    .accident-marker-wave {
        position: absolute;
        top: 50%;
        left: 50%;
        width: 34px;
        height: 34px;
        transform: translate(-50%, -50%);
        border-radius: 50%;
        border: 2px solid rgba(217, 83, 79, 0.7);
        opacity: 0;
        animation: accident-wave 2.4s infinite;
    }
    .accident-marker-wave.wave-delay {
        animation-delay: 1.2s;
    }
    @keyframes accident-wave {
        0% {
            opacity: 0.7;
            transform: translate(-50%, -50%) scale(0.3);
        }
        60% {
            opacity: 0;
            transform: translate(-50%, -50%) scale(1.6);
        }
        100% {
            opacity: 0;
            transform: translate(-50%, -50%) scale(1.6);
        }
    }
    @keyframes accident-core-pulse {
        0%, 100% {
            transform: translate(-50%, -50%) scale(1);
            box-shadow: 0 0 10px rgba(217, 83, 79, 0.7);
        }
        50% {
            transform: translate(-50%, -50%) scale(1.15);
            box-shadow: 0 0 18px rgba(217, 83, 79, 0.9);
        }
    }
    </style>
    """
    m.get_root().header.add_child(folium.Element(accident_marker_css))

    accident_marker_html = """
    <div class="accident-marker">
        <div class="accident-marker-wave"></div>
        <div class="accident-marker-wave wave-delay"></div>
        <div class="accident-marker-core"><i class="fa fa-exclamation"></i></div>
    </div>
    """

    # Animated vehicle marker styling for smoother motion cues
    vehicle_marker_css = """
    <style>
    .leaflet-marker-icon.vehicle1-marker,
    .leaflet-marker-icon.vehicle2-marker {
        transition: transform 0.25s linear;
    }
    .vehicle-marker {
        position: relative;
        width: 32px;
        height: 32px;
        display: flex;
        align-items: center;
        justify-content: center;
    }
    .vehicle-marker-primary .vehicle-body {
        background: #d9534f;
        box-shadow: none;
    }
    .vehicle-marker-secondary .vehicle-body {
        background: #5cb85c;
        box-shadow: none;
    }
    .vehicle-marker .vehicle-body {
        position: relative;
        width: 22px;
        height: 22px;
    border-radius: 4px;
    color: #ffffff;
    display: flex;
    align-items: center;
    justify-content: center;
    border: 2px solid rgba(255, 255, 255, 0.9);
    }
    .vehicle-marker .vehicle-body i {
        font-size: 13px;
    }
    .vehicle-speedline {
        position: absolute;
        left: 4px;
        width: 10px;
        height: 2px;
        background: rgba(255, 255, 255, 0.35);
        border-radius: 1px;
        opacity: 0;
        animation: vehicle-trail 0.6s linear infinite;
    }
    .vehicle-marker .vehicle-speedline:nth-of-type(1) { top: 11px; animation-delay: 0.1s; }
    .vehicle-marker .vehicle-speedline:nth-of-type(2) { top: 18px; animation-delay: 0.3s; }
    .vehicle-marker.impact-pulse {
        animation: vehicle-impact-pulse 0.6s ease-in-out 1;
    }
    @keyframes vehicle-impact-pulse {
        0%, 100% {
            transform: scale(1);
            box-shadow: 0 0 10px rgba(255, 255, 255, 0.0);
        }
        40% {
            transform: scale(1.18);
            box-shadow: 0 0 16px rgba(255, 255, 255, 0.45);
        }
        70% {
            transform: scale(0.93);
            box-shadow: 0 0 6px rgba(255, 255, 255, 0.25);
        }
    }
    @keyframes vehicle-trail {
        0% { transform: translateX(0); opacity: 0; }
        30% { opacity: 0.85; }
        100% { transform: translateX(-6px); opacity: 0; }
    }
    .collision-effect-icon {
        pointer-events: none;
    }
    .collision-burst {
        position: relative;
        width: 48px;
        height: 48px;
    }
    .collision-core {
        position: absolute;
        top: 50%;
        left: 50%;
        width: 18px;
        height: 18px;
        transform: translate(-50%, -50%);
        background: radial-gradient(circle, #ffedb5 0%, #f0ad4e 55%, #d9534f 100%);
        border-radius: 50%;
        display: flex;
        align-items: center;
        justify-content: center;
        color: #8a1f1f;
        box-shadow: 0 0 14px rgba(217, 83, 79, 0.8);
        animation: collision-core-flash 0.9s ease-out 1;
    }
    .collision-ring {
        position: absolute;
        top: 50%;
        left: 50%;
        width: 20px;
        height: 20px;
        border-radius: 50%;
        border: 2px solid rgba(217, 83, 79, 0.65);
        transform: translate(-50%, -50%) scale(0.2);
        opacity: 0;
        animation: collision-ring-expand 1.6s ease-out infinite;
    }
    .collision-ring.ring-delay {
        animation-delay: 0.5s;
    }
    @keyframes collision-ring-expand {
        0% {
            transform: translate(-50%, -50%) scale(0.2);
            opacity: 0.75;
        }
        70% {
            transform: translate(-50%, -50%) scale(1.8);
            opacity: 0;
        }
        100% {
            transform: translate(-50%, -50%) scale(1.8);
            opacity: 0;
        }
    }
    @keyframes collision-core-flash {
        0% { transform: translate(-50%, -50%) scale(0.6); opacity: 0.0; }
        30% { transform: translate(-50%, -50%) scale(1.2); opacity: 1; }
        100% { transform: translate(-50%, -50%) scale(1); opacity: 0; }
    }
    </style>
    """
    m.get_root().header.add_child(folium.Element(vehicle_marker_css))

    # Approximate conversion from degrees to meters (fallback usage)
    meters_per_degree = 111320 

    road_line: LineString | None = None
    road_length_meters: float | None = None

    # 2. Define the road segment for the simulation
    if roads_gdf_latlon is not None and not roads_gdf_latlon.empty:
        exploded_latlon = roads_gdf_latlon.explode(index_parts=False, ignore_index=True)
        exploded_merc = exploded_latlon.to_crs('EPSG:3857')
        candidates: list[tuple[LineString, float]] = []

        for geom_latlon, geom_merc in zip(exploded_latlon.geometry, exploded_merc.geometry):
            line_latlon: LineString | None = None
            line_merc = None

            if isinstance(geom_latlon, LineString):
                line_latlon = geom_latlon
                line_merc = geom_merc
            elif isinstance(geom_latlon, MultiLineString):
                segments_latlon = list(geom_latlon.geoms)
                segments_merc = list(geom_merc.geoms)
                if not segments_latlon:
                    continue
                idx = max(range(len(segments_latlon)), key=lambda i: segments_merc[i].length)
                line_latlon = segments_latlon[idx]
                line_merc = segments_merc[idx]

            if line_latlon is None or line_merc is None:
                continue

            length_m = float(line_merc.length)
            if length_m < 40:  # skip extremely short segments
                continue
            candidates.append((line_latlon, length_m))

        if candidates:
            road_line, road_length_meters = random.choice(candidates)
            road_coords = [(coord[1], coord[0]) for coord in road_line.coords]
            folium.PolyLine(
                locations=road_coords,
                color='blue',
                weight=5,
                opacity=0.8,
                popup=f"Labo Road Segment (Length: {road_length_meters:.0f}m)"
            ).add_to(m)
            print(f"Using GIS Labo road segment for simulation (length: {road_length_meters:.0f} meters)")
        else:
            print("No suitable GIS road segments found; attempting OSM fallback.")

    if road_line is None and streets is not None and not streets.empty:
        streets_projected = streets.to_crs('EPSG:3857')
        streets_projected['length'] = streets_projected.geometry.length

        labo_streets = streets[
            (streets.geometry.bounds.miny >= 14.1) &
            (streets.geometry.bounds.maxy <= 14.25) &
            (streets.geometry.bounds.minx >= 122.7) &
            (streets.geometry.bounds.maxx <= 122.95)
        ].copy()

        labo_indices = labo_streets.index
        labo_streets_projected = streets_projected.loc[labo_indices]
        labo_streets['length'] = labo_streets_projected['length']
        suitable_roads = labo_streets[labo_streets['length'] > 50].copy()

        if not suitable_roads.empty:
            selected_road = suitable_roads.sample(n=1).iloc[0]
            road_geom = selected_road.geometry

            if isinstance(road_geom, LineString):
                road_line = road_geom
            elif hasattr(road_geom, '__len__') and len(road_geom) > 0:
                road_line = road_geom[0]
            else:
                road_line = None

            if road_line is not None:
                road_length_meters = float(selected_road['length'])
                road_coords = [(coord[1], coord[0]) for coord in road_line.coords]
                folium.PolyLine(locations=road_coords, color='blue', weight=5, opacity=0.8,
                                popup=f"Random Labo Road (Length: {road_length_meters:.0f}m)").add_to(m)
                print(f"Using randomly selected Labo road from OSM (length: {road_length_meters:.0f} meters)")

        if road_line is None:
            broader_suitable_roads = streets_projected[streets_projected['length'] > 100].copy()
            broader_indices = broader_suitable_roads.index
            broader_suitable_roads_orig = streets.loc[broader_indices].copy()
            broader_suitable_roads_orig['length'] = broader_suitable_roads['length']

            if not broader_suitable_roads_orig.empty:
                selected_road = broader_suitable_roads_orig.sample(n=1).iloc[0]
                road_geom = selected_road.geometry

                if isinstance(road_geom, LineString):
                    road_line = road_geom
                elif hasattr(road_geom, '__len__') and len(road_geom) > 0:
                    road_line = road_geom[0]
                else:
                    road_line = None

                if road_line is not None:
                    road_length_meters = float(selected_road['length'])
                    road_coords = [(coord[1], coord[0]) for coord in road_line.coords]
                    folium.PolyLine(locations=road_coords, color='blue', weight=5, opacity=0.8,
                                    popup=f"Random Road (Broader Area, Length: {road_length_meters:.0f}m)").add_to(m)
                    print(f"Using randomly selected road from broader OSM area (length: {road_length_meters:.0f} meters)")

    if road_line is None or road_length_meters is None or road_length_meters <= 0:
        # Final fallback to straight line for animation continuity
        start_point = (map_center[0], map_center[1] - 0.01)
        end_point = (map_center[0], map_center[1] + 0.01)
        road_line = LineString([start_point, end_point])
        road_length_degrees = road_line.length
        road_length_meters = float(road_length_degrees * meters_per_degree)
        folium.PolyLine(locations=[start_point, end_point], color='blue', weight=5, opacity=0.8,
                        popup="Simulation Road (Fallback)").add_to(m)
        print("Using fallback straight line road for simulation")

    # 3. Prepare data for custom JavaScript animation (more reliable than TimestampedGeoJson)
    # Subsample data to reduce number of frames for better performance
    subsample_rate = 10  # Every 10th frame (0.1s intervals)
    sim_data_subsampled = sim_data.iloc[::subsample_rate].copy()

    # Extract vehicle positions for JavaScript animation
    vehicle1_positions = []
    vehicle2_positions = []
    pedestrian_positions = []
    time_points = []

    for _, row in sim_data_subsampled.iterrows():
        time_points.append(row['time'])

        # Vehicle 1 position
        fraction_of_road = row[f'position_{vehicle1_name.lower().replace(" ", "_")}'] / road_length_meters
        if fraction_of_road > 1.0: fraction_of_road = 1.0
        point_on_line = road_line.interpolate(fraction_of_road, normalized=True)
        vehicle1_positions.append([point_on_line.y, point_on_line.x])  # [lat, lon]

        # Vehicle 2 or pedestrian position
        if vehicle2_name:
            fraction_of_road = row[f'position_{vehicle2_name.lower().replace(" ", "_")}'] / road_length_meters
            if fraction_of_road > 1.0: fraction_of_road = 1.0
            point_on_line = road_line.interpolate(fraction_of_road, normalized=True)
            vehicle2_positions.append([point_on_line.y, point_on_line.x])
            pedestrian_positions.append(None)
        elif pedestrian_name:
            fraction_of_road = row[f'position_{pedestrian_name.lower().replace(" ", "_")}'] / road_length_meters
            if fraction_of_road > 1.0: fraction_of_road = 1.0
            point_on_line = road_line.interpolate(fraction_of_road, normalized=True)
            pedestrian_positions.append([point_on_line.y, point_on_line.x])
            vehicle2_positions.append(None)
        else:
            vehicle2_positions.append(None)
            pedestrian_positions.append(None)

    collision_frame_index = None
    if collision_time is not None and time_points:
        for idx, t_val in enumerate(time_points):
            if t_val >= collision_time:
                collision_frame_index = idx
                break
        if collision_frame_index is None:
            collision_frame_index = len(time_points) - 1

    accident_marker_coords = map_center
    if collision_frame_index is not None:
        def _pick_position(position_list):
            if not position_list:
                return None
            idx = min(collision_frame_index, len(position_list) - 1)
            candidate = position_list[idx]
            return candidate if candidate else None

        chosen = _pick_position(vehicle1_positions)
        if chosen is None:
            chosen = _pick_position(vehicle2_positions)
        if chosen is None:
            chosen = _pick_position(pedestrian_positions)
        if chosen is not None:
            accident_marker_coords = chosen

    folium.Marker(
        location=accident_marker_coords,
        popup=f"Accident Location: {location}",
        icon=folium.DivIcon(html=accident_marker_html, icon_size=(34, 34), icon_anchor=(17, 17), class_name='')
    ).add_to(m)

    # 4. Add custom JavaScript animation instead of TimestampedGeoJson
    # Create vehicle markers
    vehicle1_marker = None
    vehicle2_marker = None
    pedestrian_marker = None

    if vehicle1_positions:
        vehicle1_icon_html = """
        <div class="vehicle-marker vehicle-marker-primary">
            <div class="vehicle-body"><i class="fa fa-car"></i></div>
            <span class="vehicle-speedline"></span>
            <span class="vehicle-speedline"></span>
            <span class="vehicle-speedline"></span>
        </div>
        """
        vehicle1_marker = folium.Marker(
            location=vehicle1_positions[0],
            popup=f"{vehicle1_name}<br>Initial position",
            icon=folium.DivIcon(html=vehicle1_icon_html, icon_size=(32, 32), icon_anchor=(16, 16), class_name='vehicle1-marker')
        ).add_to(m)

    if vehicle2_name and vehicle2_positions and vehicle2_positions[0]:
        vehicle2_icon_class = 'fa-truck' if 'truck' in vehicle2_name.lower() else 'fa-bus' if 'bus' in vehicle2_name.lower() else 'fa-car'
        vehicle2_icon_html = f"""
        <div class=\"vehicle-marker vehicle-marker-secondary\">
            <div class=\"vehicle-body\"><i class=\"fa {vehicle2_icon_class}\"></i></div>
            <span class=\"vehicle-speedline\"></span>
            <span class=\"vehicle-speedline\"></span>
            <span class=\"vehicle-speedline\"></span>
        </div>
        """
        vehicle2_marker = folium.Marker(
            location=vehicle2_positions[0],
            popup=f"{vehicle2_name}<br>Initial position",
            icon=folium.DivIcon(html=vehicle2_icon_html, icon_size=(32, 32), icon_anchor=(16, 16), class_name='vehicle2-marker')
        ).add_to(m)
    elif pedestrian_name and pedestrian_positions and pedestrian_positions[0]:
        pedestrian_marker = folium.Marker(
            location=pedestrian_positions[0],
            popup=f"{pedestrian_name}<br>Initial position",
            icon=folium.Icon(color='purple', icon='male', prefix='fa')
        ).add_to(m)

    vehicle1_marker_js = vehicle1_marker.get_name() if vehicle1_marker else 'null'
    vehicle2_marker_js = vehicle2_marker.get_name() if vehicle2_marker else 'null'
    pedestrian_marker_js = pedestrian_marker.get_name() if pedestrian_marker else 'null'

    # Add custom JavaScript for animation
    animation_js = f"""
    <script>
    // Animation data
    var vehicle1Positions = {vehicle1_positions};
    var vehicle2Positions = {vehicle2_positions};
    var pedestrianPositions = {pedestrian_positions};
    var timePoints = {time_points};
    var collisionFrameIndex = {'null' if collision_frame_index is None else collision_frame_index};
    var collisionTriggered = false;
    var currentFrame = 0;
    var isPlaying = false;
    var animationSpeed = 200; // milliseconds between frames
    var animationInterval;

    // Get marker objects from Folium
    var vehicle1Marker = {vehicle1_marker_js};
    var vehicle2Marker = {vehicle2_marker_js};
    var pedestrianMarker = {pedestrian_marker_js};
    var animationMap = null;
    var collisionMarker = null;

    function resolveAnimationMap() {{
        if (vehicle1Marker && vehicle1Marker._map) {{
            animationMap = vehicle1Marker._map;
        }} else if (vehicle2Marker && vehicle2Marker._map) {{
            animationMap = vehicle2Marker._map;
        }} else if (pedestrianMarker && pedestrianMarker._map) {{
            animationMap = pedestrianMarker._map;
        }}
    }}

    function initializeAnimation() {{
        if (!vehicle1Positions.length) {{
            console.warn('No vehicle position data available for animation.');
            return;
        }}

        resolveAnimationMap();
        if (!animationMap) {{
            // Try again shortly if map not yet attached
            setTimeout(initializeAnimation, 250);
            return;
        }}

        updateAnimationFrame();
    }}

    setTimeout(initializeAnimation, 500);

    function updateAnimationFrame() {{
        if (currentFrame >= vehicle1Positions.length) {{
            currentFrame = 0; // Loop back to start
            collisionTriggered = false;
            removeCollisionEffect();
        }}

        // Update vehicle1 position
        if (vehicle1Marker) {{
            vehicle1Marker.setLatLng(vehicle1Positions[currentFrame]);
            vehicle1Marker.setPopupContent("{vehicle1_name}<br>Time: " + timePoints[currentFrame].toFixed(2) + "s");
        }}

        // Update vehicle2 or pedestrian position
        if (vehicle2Marker) {{
            if (vehicle2Positions[currentFrame]) {{
                vehicle2Marker.setLatLng(vehicle2Positions[currentFrame]);
                vehicle2Marker.setPopupContent("{vehicle2_name if vehicle2_name else pedestrian_name}<br>Time: " + timePoints[currentFrame].toFixed(2) + "s");
            }}
        }}

        if (pedestrianMarker) {{
            if (pedestrianPositions[currentFrame]) {{
                pedestrianMarker.setLatLng(pedestrianPositions[currentFrame]);
                pedestrianMarker.setPopupContent("{pedestrian_name if pedestrian_name else 'Pedestrian'}<br>Time: " + timePoints[currentFrame].toFixed(2) + "s");
            }}
        }}

        if (!collisionTriggered && collisionFrameIndex !== null && currentFrame >= collisionFrameIndex) {{
            collisionTriggered = true;
            triggerCollisionEffect(vehicle1Positions[currentFrame]);
            applyImpactAnimation(vehicle1Marker);
            applyImpactAnimation(vehicle2Marker);
            applyImpactAnimation(pedestrianMarker);
        }}

        currentFrame++;
    }}

    function playAnimation() {{
        if (!isPlaying) {{
            isPlaying = true;
            animationInterval = setInterval(updateAnimationFrame, animationSpeed);
            document.getElementById('playBtn').innerHTML = '<i class="fa fa-pause"></i> Pause';
            document.getElementById('playBtn').className = 'btn btn-warning btn-sm';
        }} else {{
            pauseAnimation();
        }}
    }}

    function pauseAnimation() {{
        isPlaying = false;
        clearInterval(animationInterval);
        document.getElementById('playBtn').innerHTML = '<i class="fa fa-play"></i> Play';
        document.getElementById('playBtn').className = 'btn btn-success btn-sm';
    }}

    function resetAnimation() {{
        pauseAnimation();
        currentFrame = 0;
        collisionTriggered = false;
        removeCollisionEffect();
        updateAnimationFrame();
    }}

    function changeSpeed(speed) {{
        animationSpeed = speed;
        if (isPlaying) {{
            clearInterval(animationInterval);
            animationInterval = setInterval(updateAnimationFrame, animationSpeed);
        }}
    }}

    // Add control panel to map
    var controlPanel = L.control({{position: 'topright'}});
    controlPanel.onAdd = function(map) {{
        var div = L.DomUtil.create('div', 'animation-controls');
        div.innerHTML = `
            <div style="background: white; padding: 10px; border-radius: 5px; box-shadow: 0 0 10px rgba(0,0,0,0.2);">
                <h4 style="margin: 0 0 10px 0; font-size: 14px;">Animation Controls</h4>
                <button id="playBtn" class="btn btn-success btn-sm" onclick="playAnimation()">
                    <i class="fa fa-play"></i> Play
                </button>
                <button class="btn btn-secondary btn-sm" onclick="resetAnimation()">
                    <i class="fa fa-refresh"></i> Reset
                </button><br><br>
                <label style="font-size: 12px;">Speed:</label><br>
                <input type="range" min="50" max="1000" value="200" step="50"
                       onchange="changeSpeed(this.value)" style="width: 100px;">
                <span id="speedValue" style="font-size: 12px;">200ms</span>
            </div>
        `;
        return div;
    }};
    controlPanel.addTo(map);

    // Update speed display
    document.addEventListener('input', function(e) {{
        if (e.target.type === 'range') {{
            document.getElementById('speedValue').textContent = e.target.value + 'ms';
        }}
    }});

    function triggerCollisionEffect(position) {{
        if (!animationMap || !position) {{
            return;
        }}
        removeCollisionEffect();
        var collisionHtml = `
            <div class="collision-burst">
                <span class="collision-ring"></span>
                <span class="collision-ring ring-delay"></span>
                <div class="collision-core"><i class="fa fa-bolt"></i></div>
            </div>`;
        collisionMarker = L.marker(position, {{
            icon: L.divIcon({{
                html: collisionHtml,
                className: 'collision-effect-icon',
                iconSize: [48, 48],
                iconAnchor: [24, 24]
            }}),
            interactive: false
        }}).addTo(animationMap);
        setTimeout(function() {{
            removeCollisionEffect();
        }}, 2400);
    }}

    function removeCollisionEffect() {{
        if (collisionMarker && animationMap) {{
            animationMap.removeLayer(collisionMarker);
            collisionMarker = null;
        }}
    }}

    function applyImpactAnimation(marker) {{
        if (!marker || !marker.getElement) {{
            return;
        }}
        var element = marker.getElement();
        if (!element) {{
            return;
        }}
        var inner = element.querySelector('.vehicle-marker');
        if (!inner) {{
            return;
        }}
        inner.classList.remove('impact-pulse');
        void inner.offsetWidth;
        inner.classList.add('impact-pulse');
        setTimeout(function() {{
            inner.classList.remove('impact-pulse');
        }}, 650);
    }}
    </script>
    """

    # Add the JavaScript to the map
    m.get_root().html.add_child(folium.Element(animation_js))

    # 5. Add historical accident data markers
    try:
        historical_data = pd.read_csv('accident_data.csv')
        # Group by location and count accidents
        location_counts = historical_data.groupby('location').size().reset_index(name='accident_count')
        severity_counts = historical_data.groupby(['location', 'severity']).size().unstack(fill_value=0)
        
        # Accurate coordinates for Labo locations (based on real geography)
        location_coords = {
            'Bayabas': [14.158, 122.825],  # Northeast area
            'Main Highway': [14.156, 122.83],  # Central highway
            'Barangay 1': [14.152, 122.835],  # Southwest area
            'Barangay 2': [14.154, 122.828],  # South area
            'Barangay 3': [14.160, 122.832],  # North area
            'Barangay 4': [14.158, 122.838],  # East area
            'Barangay 5': [14.162, 122.826],  # Northeast area
        }
        
        for _, row in location_counts.iterrows():
            loc_name = row['location']
            count = row['accident_count']
            
            # Use accurate coordinates if available, otherwise place near center
            if loc_name in location_coords:
                marker_lat, marker_lon = location_coords[loc_name]
            else:
                # Fallback: place near center with small offset
                offset_lat = random.uniform(-0.002, 0.002)
                offset_lon = random.uniform(-0.002, 0.002)
                marker_lat = map_center[0] + offset_lat
                marker_lon = map_center[1] + offset_lon
            
            # Get severity breakdown for this location
            severities = severity_counts.loc[loc_name] if loc_name in severity_counts.index else {}
            severity_text = ", ".join([f"{sev}: {cnt}" for sev, cnt in severities.items() if cnt > 0])
            
            # Color based on total accidents
            if count >= 5:
                color = 'red'
            elif count >= 3:
                color = 'orange'
            else:
                color = 'blue'
            
            folium.Marker(
                location=[marker_lat, marker_lon],
                popup=f"<b>{loc_name}</b><br>Total Accidents: {count}<br>Severity: {severity_text}",
                icon=folium.Icon(color=color, icon='exclamation-triangle', prefix='fa')
            ).add_to(m)
        
        print(f"Added {len(location_counts)} historical accident location markers")
    except Exception as e:
        print(f"Could not load historical accident data: {e}")

    # 6. Add some basic landmarks or features (example)
    # Add a sample traffic signal marker
    folium.Marker(
        location=[map_center[0] + 0.002, map_center[1] + 0.002],
        popup="Traffic Signal",
        icon=folium.Icon(color='orange', icon='traffic-light', prefix='fa')
    ).add_to(m)

    # 6. Add legend summarizing map symbology derived from GIS overlay and markers
    legend_html = """
    {% macro html(this, kwargs) %}
    <div style="position: fixed; bottom: 20px; left: 20px; z-index: 9999; background-color: white; border: 1px solid #ccc; border-radius: 6px; box-shadow: 0 2px 6px rgba(0,0,0,0.25); padding: 12px 14px; font-size: 13px; line-height: 1.5; min-width: 220px;">
        <div style="font-weight: 600; margin-bottom: 8px;">Map Legend</div>
        <div style="display: flex; align-items: center; margin-bottom: 6px;">
            <span style="display: inline-block; width: 26px; height: 4px; background-color: #ff7f0e; margin-right: 8px;"></span>
            <span>Labo road network (GIS overlay)</span>
        </div>
        <div style="display: flex; align-items: center; margin-bottom: 6px;">
            <span style="display: inline-block; width: 26px; height: 4px; background-color: #1f77b4; margin-right: 8px;"></span>
            <span>Simulated travel path</span>
        </div>
        <div style="display: flex; align-items: center; margin-bottom: 6px;">
            <i class="fa fa-exclamation-triangle" style="color: #d9534f; margin-right: 8px;"></i>
            <span>High accident density (≥5)</span>
        </div>
        <div style="display: flex; align-items: center; margin-bottom: 6px;">
            <i class="fa fa-exclamation-triangle" style="color: #f0ad4e; margin-right: 8px;"></i>
            <span>Moderate accident density (3-4)</span>
        </div>
        <div style="display: flex; align-items: center; margin-bottom: 6px;">
            <i class="fa fa-exclamation-triangle" style="color: #0275d8; margin-right: 8px;"></i>
            <span>Low accident density (&lt;3)</span>
        </div>
        <div style="display: flex; align-items: center; margin-bottom: 6px;">
            <i class="fa fa-car" style="color: #d9534f; margin-right: 8px;"></i>
            <span>Vehicle 1 (simulated)</span>
        </div>
        <div style="display: flex; align-items: center;">
            <i class="fa fa-truck" style="color: #5cb85c; margin-right: 8px;"></i>
            <span>Vehicle 2 / other entities</span>
        </div>
        <div style="display: flex; align-items: center; margin-top: 6px;">
            <i class="fa fa-traffic-light" style="color: #f0ad4e; margin-right: 8px;"></i>
            <span>Annotated traffic signal / landmark</span>
        </div>
    </div>
    {% endmacro %}
    """

    legend = MacroElement()
    legend._template = Template(legend_html)
    m.get_root().add_child(legend)

    # 7. Save to an HTML file
    output_filename = 'road_accident_map_simulation.html'
    m.save(output_filename)
    print(f"\nMap visualization saved as '{output_filename}'. Open this file in a web browser.")


# Example simulation
if __name__ == "__main__":
    # Randomize accident type
    accident_type = random.choice(list(accident_types.keys()))
    print(f"Simulating accident type: {accident_type}")
    
    # Sample parameters from historical data sources (PDF Materials and Methods A. Data Sources)
    cause, location, primary_vehicle_type, road_condition, lighting_condition, weather_condition, human_factor = sample_from_data_sources()
    print(f"Based on data sources: Location={location}, Cause={cause}, Road={road_condition}, Lighting={lighting_condition}, Weather={weather_condition}, Human Factor={human_factor}")

    environment = build_environment_state(location, road_condition, lighting_condition, weather_condition)
    driver_profile = DriverProfile.from_factor(human_factor)
    intervention_plan = suggest_interventions(location, cause, environment.road_condition, environment.lighting_condition, human_factor)

    vehicle1_spec, vehicle2_spec, pedestrian_spec, angle_of_impact = prepare_accident_entities(accident_type, primary_vehicle_type)
    baseline_vehicle1, baseline_vehicle2, baseline_pedestrian = instantiate_entities(vehicle1_spec, vehicle2_spec, pedestrian_spec)

    initial_v1 = vehicle1_spec["velocity"] if vehicle1_spec else None
    initial_v2 = vehicle2_spec["velocity"] if vehicle2_spec else None

    print_simulation_steps(
        accident_type,
        cause,
        location,
        environment.road_condition,
        environment.lighting_condition,
        baseline_vehicle1,
        baseline_vehicle2,
        baseline_pedestrian,
        weather=environment.weather_condition,
        driver_profile=driver_profile,
        environment=environment,
        interventions=intervention_plan
    )

    # Baseline simulation without interventions
    pos1, pos2, vel1, vel2, sim_time, impact_force, severity, risk_score, sim_data = simulate_collision(
        baseline_vehicle1,
        baseline_vehicle2,
        pedestrian=baseline_pedestrian,
        accident_type=accident_type,
        road_condition=environment.road_condition,
        lighting_condition=environment.lighting_condition,
        angle_of_impact=angle_of_impact,
        environment=environment,
        driver_profile=driver_profile,
        interventions=None,
        verbose=True
    )

    print("Simulation Data (baseline):")
    print(sim_data.head())

    # Intervention simulation for what-if comparison
    if intervention_plan.interventions:
        intervention_vehicle1, intervention_vehicle2, intervention_pedestrian = instantiate_entities(vehicle1_spec, vehicle2_spec, pedestrian_spec)
        _, _, _, _, sim_time_int, impact_force_int, severity_int, risk_score_int, _ = simulate_collision(
            intervention_vehicle1,
            intervention_vehicle2,
            pedestrian=intervention_pedestrian,
            accident_type=accident_type,
            road_condition=environment.road_condition,
            lighting_condition=environment.lighting_condition,
            angle_of_impact=angle_of_impact,
            environment=environment,
            driver_profile=driver_profile,
            interventions=intervention_plan.interventions,
            verbose=False
        )
    else:
        severity_int = severity
        impact_force_int = impact_force
        risk_score_int = risk_score
        sim_time_int = sim_time

    comparison_payload = {
        "baseline": {
            "severity": severity,
            "impact_force": impact_force,
            "risk_score": risk_score,
            "collision_time": sim_time
        },
        "intervention": {
            "severity": severity_int,
            "impact_force": impact_force_int,
            "risk_score": risk_score_int,
            "collision_time": sim_time_int
        }
    } if intervention_plan.interventions else None

    validation_text = validate_with_historical(location, severity)

    report = generate_report(
        baseline_vehicle1,
        baseline_vehicle2,
        baseline_pedestrian,
        accident_type,
        sim_time,
        impact_force,
        severity,
        environment.road_condition,
        environment.lighting_condition,
        initial_v1,
        initial_v2,
        cause,
        location,
        angle_of_impact=angle_of_impact,
        weather=environment.weather_condition,
        driver_profile=driver_profile,
        environment=environment,
        risk_score=risk_score,
        intervention_plan=intervention_plan,
        validation_text=validation_text,
        comparison=comparison_payload
    )
    print(report)

    # Monte Carlo analysis offers a quick predictive snapshot for planning
    monte_carlo_summary = run_monte_carlo_intervention_analysis(runs=20)
    if monte_carlo_summary:
        print(f"Monte Carlo comparison (baseline vs interventions): {monte_carlo_summary}")

    geojson_path = create_labo_roads_geojson()
    if geojson_path is None:
        print("Warning: Could not prepare GIS/labo_roads.geojson. The map overlay may be missing.")
    else:
        print(f"Prepared road overlay at {geojson_path}")

    create_map_visualization(
        sim_data,
        vehicle1_name=baseline_vehicle1.name if baseline_vehicle1 else "Vehicle 1",
        vehicle2_name=baseline_vehicle2.name if baseline_vehicle2 else None,
        pedestrian_name=baseline_pedestrian.name if baseline_pedestrian else None,
        location=location,
        labo_geojson_path=geojson_path,
        collision_time=sim_time
    )

    # Persist results for longitudinal analysis (append rather than overwrite)
    if severity != "none":
        try:
            from datetime import datetime

            new_accident = {
                'cause': cause,
                'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'location': location,
                'vehicle_type': primary_vehicle_type,
                'weather': environment.weather_condition,
                'lighting': 'Daylight' if environment.lighting_condition == 'good' else 'Night',
                'road_characteristics': {
                    'dry': 'Dry road',
                    'wet': 'Wet road',
                    'slippery': 'Slippery surface'
                }.get(environment.road_condition, environment.road_condition),
                'human_factors': human_factor,
                'severity': severity.capitalize()
            }

            try:
                historical_df = pd.read_csv('accident_data.csv')
                historical_df = pd.concat([historical_df, pd.DataFrame([new_accident])], ignore_index=True)
            except FileNotFoundError:
                historical_df = pd.DataFrame([new_accident])

            historical_df.to_csv('accident_data.csv', index=False)
            print(f"Accident data appended to accident_data.csv: {severity.capitalize()} accident in {location}")
        except Exception as e:
            print(f"Could not save accident data to CSV: {e}")
    else:
        print("No collision occurred - accident not saved to database")

    # Plot positions
    time_array = np.arange(0, len(pos1)) * 0.01
    plt.plot(time_array, pos1, label=baseline_vehicle1.name if baseline_vehicle1 else "Vehicle")
    if baseline_vehicle2:
        plt.plot(time_array, pos2, label=baseline_vehicle2.name)
    elif baseline_pedestrian:
        plt.plot(time_array, pos2, label=baseline_pedestrian.name)
    plt.xlabel('Time (s)')
    plt.ylabel('Position (m)')
    plt.legend()
    plt.title('Entity Positions During Simulation')
    # plt.show()  # Commented out to avoid blocking