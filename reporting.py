# reporting.py
# Reporting and visualization functions for road accident simulation

import matplotlib.pyplot as plt
import numpy as np
from typing import Optional, Dict, Any, List
import os
import json
import random
from pathlib import Path
from datetime import datetime, timedelta

# GIS and mapping libraries
try:
    import geopandas as gpd
    from shapely.geometry import LineString, MultiLineString
    import osmnx as ox
    import folium
    from branca.element import MacroElement
    from jinja2 import Template
    import pandas as pd
except ImportError as e:
    print(f"Warning: Some GIS/mapping libraries not available: {e}")

from constants import accident_types, LABO_COORDINATES, LABO_BARANGAYS
from utils import generate_recommendations, generate_systemic_summary, validate_with_historical, rng_for_location
from models import EnvironmentState, DriverProfile, InterventionPlan

def parse_labo_roads_geojson(geojson_path: str = "GIS/labo_roads.geojson"):
    try:
        import json
        with open(geojson_path, 'r', encoding='utf-8') as f:
            geojson_data = json.load(f)

        roads = []
        for feature in geojson_data.get('features', []):
            if feature.get('geometry', {}).get('type') == 'LineString':
                coordinates = feature['geometry']['coordinates']
                # Convert [lng, lat] to [lat, lng] for folium
                road_coords = [[coord[1], coord[0]] for coord in coordinates]
                roads.append({
                    'coordinates': road_coords,
                    'properties': feature.get('properties', {})
                })

        print(f"Parsed {len(roads)} road segments from GeoJSON")
        return roads
    except Exception as e:
        print(f"Error parsing GeoJSON: {e}")
        return []

def latlng_to_pixel(lat, lng, bounds, map_width=800, map_height=600):
    """Convert lat/lng coordinates to pixel positions relative to map bounds."""
    min_lat, max_lat = bounds['lat']
    min_lng, max_lng = bounds['lng']

    # Calculate relative position within bounds
    lat_ratio = (lat - min_lat) / (max_lat - min_lat)
    lng_ratio = (lng - min_lng) / (max_lng - min_lng)

    # Convert to pixel coordinates (flip Y axis since lat increases upward)
    x = lng_ratio * map_width
    y = (1 - lat_ratio) * map_height

    return x, y

def get_map_bounds_from_roads(roads):
    """Calculate map bounds from road coordinates."""
    if not roads:
        return {'lat': [14.08, 14.11], 'lng': [122.58, 122.65]}  # Default Labo bounds

    all_coords = []
    for road in roads:
        all_coords.extend(road['coordinates'])

    lats = [coord[0] for coord in all_coords]
    lngs = [coord[1] for coord in all_coords]

    return {
        'lat': [min(lats), max(lats)],
        'lng': [min(lngs), max(lngs)]
    }

def print_simulation_steps(accident_type: str, cause: str, location: str, road_condition: str, lighting_condition: str, vehicle1: Optional[Any], vehicle2: Optional[Any], pedestrian: Optional[Any], weather: Optional[str], driver_profile: Optional[DriverProfile], environment: Optional[EnvironmentState], interventions: Optional[InterventionPlan]):
    """Print step-by-step breakdown of the road accident simulation process"""
    print("\n" + "="*80)
    print("ROAD ACCIDENT SIMULATION STEPS")
    print("="*80)
    
    print(f"\nStep 1: Accident Scenario Setup")
    print(f"   - Accident Type: {accident_types[accident_type]['description']}")
    print(f"   - Location: {location}, Labo, Camarines Norte")
    print(f"   - Primary Cause: {cause}")
    print(f"   - Road Conditions: {road_condition}")
    print(f"   - Lighting Conditions: {lighting_condition}")
    if weather:
        print(f"   - Weather: {weather}")
    if environment:
        print(f"   - Road Geometry: {environment.curvature}, Approx. Slope: {environment.slope:.1f}°")
    if driver_profile:
        print(f"   - Driver State: {driver_profile.factor} ({driver_profile.notes})")
    
    print(f"\nStep 2: Vehicle/Pedestrian Initialization")
    if vehicle1:
        speed_kmh = np.linalg.norm(vehicle1.velocity) * 3.6
        print(f"   - Vehicle 1: {vehicle1.name} (Mass: {vehicle1.mass} kg, Initial Speed: {speed_kmh:.1f} km/h)")
    if vehicle2:
        speed_kmh = np.linalg.norm(vehicle2.velocity) * 3.6
        print(f"   - Vehicle 2: {vehicle2.name} (Mass: {vehicle2.mass} kg, Initial Speed: {speed_kmh:.1f} km/h)")
    if pedestrian:
        pos = getattr(pedestrian, 'position', None)
        # Format position based on whether it's a numpy array/vector or a scalar
        if pos is None:
            pos_str = 'Unknown'
        else:
            try:
                # If iterable (np.ndarray, list, tuple) and has at least two entries, print as (x, y)
                if hasattr(pos, '__len__') and len(pos) >= 2:
                    pos_str = f"({float(pos[0]):.1f}, {float(pos[1]):.1f})"
                else:
                    pos_str = f"{float(pos):.1f}"
            except Exception:
                # Fallback to string representation
                pos_str = str(pos)
        print(f"   - Pedestrian: {pedestrian.name} (Mass: {pedestrian.mass} kg, Position: {pos_str} m)")
    
    print(f"\nStep 3: Environmental Factors Applied")
    friction_map = {"dry": 0.8, "wet": 0.5, "slippery": 0.1}
    reaction_map = {"good": 0.5, "poor": 1.5}
    friction_coeff = friction_map.get(road_condition, 0.8)
    reaction_time = reaction_map.get(lighting_condition, 0.5)
    if environment:
        friction_coeff = environment.get_effective_friction(driver_profile)
        reaction_time = environment.get_effective_reaction(driver_profile)
    print(f"   - Friction Coefficient: {friction_coeff} (based on {road_condition} road)")
    print(f"   - Driver Reaction Time: {reaction_time}s (based on {lighting_condition} lighting)")
    
    if accident_type == "rear-end":
        print(f"   - Braking Applied: Vehicle 1 applies brakes after {reaction_time}s reaction time")
    
    print(f"\nStep 4: Physics Simulation Begins")
    print(f"   - Time Step: 0.01 seconds")
    print(f"   - Total Simulation Time: 10.0 seconds")
    print(f"   - Collision Detection Distance: 5.0 meters")
    
    print(f"\nStep 5: Motion Calculation")
    print(f"   - Position updates using kinematic equations:")
    print(f"     position = position + velocity * time + 0.5 * acceleration * time^2")
    print(f"     velocity = velocity + acceleration * time")
    
    if accident_type == "rear-end":
        deceleration = friction_coeff * 9.81
        print(f"   - Braking Deceleration: {deceleration:.1f} m/s² (friction * gravity)")
    
    print(f"\nStep 6: Collision Detection")
    if accident_type == "rear-end":
        print(f"   - Checking if Vehicle 1 overtakes Vehicle 2 within collision distance")
    elif accident_type == "head-on":
        print(f"   - Checking if vehicles meet within collision distance")
    elif accident_type == "side-impact":
        print(f"   - Checking perpendicular collision proximity")
    elif accident_type == "pedestrian":
        print(f"   - Checking if vehicle reaches pedestrian position")
    
    print(f"\nStep 7: Impact Analysis (if collision occurs)")
    print(f"   - Momentum Conservation: m1*v1 + m2*v2 = (m1+m2)*v_final")
    print(f"   - Impact Force Calculation: F = m * delta_v / delta_t")
    print(f"   - Severity Classification:")
    print(f"     • Minor: Impact Force < 50,000 N")
    print(f"     • Moderate: 50,000 N <= Impact Force < 150,000 N")
    print(f"     • Severe: Impact Force >= 150,000 N")
    
    print(f"\nStep 8: Data Recording")
    print(f"   - Position and velocity data collected every 0.01 seconds")
    print(f"   - Simulation data stored in pandas DataFrame")
    
    print(f"\nStep 9: Report Generation")
    print(f"   - Accident analysis and recommendations generated")
    print(f"   - Excel report created with formatted data")
    
    print(f"\nStep 10: Visualization")
    print(f"   - Position vs time plots generated for analysis")
    if interventions and interventions.interventions:
        print(f"\nStep 11: Intervention Modeling")
        from constants import intervention_effects
        for item in interventions.interventions:
            desc = intervention_effects.get(item, {}).get('description', 'Applied based on field recommendations')
            print(f"   - {item.replace('_', ' ').title()}: {desc}")
    
    print("\n" + "="*80)
    print("SIMULATION EXECUTION BEGINS")
    print("="*80 + "\n")

    if environment:
        print(f"   - Road Geometry: {environment.curvature}, Approx. Slope: {environment.slope:.1f}°")
    if driver_profile:
        print(f"   - Driver State: {driver_profile.factor} ({driver_profile.notes})")
    
    print(f"\nStep 2: Vehicle/Pedestrian Initialization")
    if vehicle1:
        speed_kmh = np.linalg.norm(vehicle1.velocity) * 3.6
        print(f"   - Vehicle 1: {vehicle1.name} (Mass: {vehicle1.mass} kg, Initial Speed: {speed_kmh:.1f} km/h)")
    if vehicle2:
        speed_kmh = np.linalg.norm(vehicle2.velocity) * 3.6
        print(f"   - Vehicle 2: {vehicle2.name} (Mass: {vehicle2.mass} kg, Initial Speed: {speed_kmh:.1f} km/h)")
    if pedestrian:
        pos = getattr(pedestrian, 'position', None)
        # Format position based on whether it's a numpy array/vector or a scalar
        if pos is None:
            pos_str = 'Unknown'
        else:
            try:
                if hasattr(pos, '__len__') and len(pos) >= 2:
                    pos_str = f"({float(pos[0]):.1f}, {float(pos[1]):.1f})"
                else:
                    pos_str = f"{float(pos):.1f}"
            except Exception:
                pos_str = str(pos)
        print(f"   - Pedestrian: {pedestrian.name} (Mass: {pedestrian.mass} kg, Position: {pos_str} m)")
    
    print(f"\nStep 3: Environmental Factors Applied")
    friction_map = {"dry": 0.8, "wet": 0.5, "slippery": 0.1}
    reaction_map = {"good": 0.5, "poor": 1.5}
    friction_coeff = friction_map.get(road_condition, 0.8)
    reaction_time = reaction_map.get(lighting_condition, 0.5)
    if environment:
        friction_coeff = environment.get_effective_friction(driver_profile)
        reaction_time = environment.get_effective_reaction(driver_profile)
    print(f"   - Friction Coefficient: {friction_coeff} (based on {road_condition} road)")
    print(f"   - Driver Reaction Time: {reaction_time}s (based on {lighting_condition} lighting)")
    
    if accident_type == "rear-end":
        print(f"   - Braking Applied: Vehicle 1 applies brakes after {reaction_time}s reaction time")
    
    print(f"\nStep 4: Physics Simulation Begins")
    print(f"   - Time Step: 0.01 seconds")
    print(f"   - Total Simulation Time: 10.0 seconds")
    print(f"   - Collision Detection Distance: 5.0 meters")
    
    print(f"\nStep 5: Motion Calculation")
    print(f"   - Position updates using kinematic equations:")
    print(f"     position = position + velocity * time + 0.5 * acceleration * time^2")
    print(f"     velocity = velocity + acceleration * time")
    
    if accident_type == "rear-end":
        deceleration = friction_coeff * 9.81
        print(f"   - Braking Deceleration: {deceleration:.1f} m/s² (friction * gravity)")
    
    print(f"\nStep 6: Collision Detection")
    if accident_type == "rear-end":
        print(f"   - Checking if Vehicle 1 overtakes Vehicle 2 within collision distance")
    elif accident_type == "head-on":
        print(f"   - Checking if vehicles meet within collision distance")
    elif accident_type == "side-impact":
        print(f"   - Checking perpendicular collision proximity")
    elif accident_type == "pedestrian":
        print(f"   - Checking if vehicle reaches pedestrian position")
    
    print(f"\nStep 7: Impact Analysis (if collision occurs)")
    print(f"   - Momentum Conservation: m1*v1 + m2*v2 = (m1+m2)*v_final")
    print(f"   - Impact Force Calculation: F = m * delta_v / delta_t")
    print(f"   - Severity Classification:")
    print(f"     • Minor: Impact Force < 50,000 N")
    print(f"     • Moderate: 50,000 N <= Impact Force < 150,000 N")
    print(f"     • Severe: Impact Force >= 150,000 N")
    
    print(f"\nStep 8: Data Recording")
    print(f"   - Position and velocity data collected every 0.01 seconds")
    print(f"   - Simulation data stored in pandas DataFrame")
    
    print(f"\nStep 9: Report Generation")
    print(f"   - Accident analysis and recommendations generated")
    print(f"   - Excel report created with formatted data")
    
    print(f"\nStep 10: Visualization")
    print(f"   - Position vs time plots generated for analysis")
    if interventions and interventions.interventions:
        print(f"\nStep 11: Intervention Modeling")
        from constants import intervention_effects
        for item in interventions.interventions:
            desc = intervention_effects.get(item, {}).get('description', 'Applied based on field recommendations')
            print(f"   - {item.replace('_', ' ').title()}: {desc}")
    
    print("\n" + "="*80)
    print("SIMULATION EXECUTION BEGINS")
    print("="*80 + "\n")


def generate_report(
    vehicle1: Optional[Any],
    vehicle2: Optional[Any],
    pedestrian: Optional[Any],
    accident_type: str,
    collision_time: float,
    impact_force: float,
    severity: str,
    road_condition: str,
    lighting: str,
    initial_v1: Optional[float],
    initial_v2: Optional[float],
    cause: str,
    location: str,
    angle_of_impact: int = 0,
    weather: Optional[str] = None,
    driver_profile: Optional[DriverProfile] = None,
    environment: Optional[EnvironmentState] = None,
    risk_score: Optional[float] = None,
    intervention_plan: Optional[InterventionPlan] = None,
    validation_text: Optional[str] = None,
    comparison: Optional[Dict] = None,
    ml_prediction: Optional[str] = None
) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import MergedCell

    accident_desc = accident_types[accident_type]["description"]
    geometry_summary = (
        f"Road Geometry: {environment.curvature} (Slope: {environment.slope:.1f}°)" if environment else "Road Geometry: Not specified"
    )
    report = f"""
Road Accident Simulation Report
================================

Accident Type: {accident_desc}
Location: {location}, Labo, Camarines Norte
Time of Accident: {collision_time:.2f} seconds into simulation
Cause: {cause}
Weather: {weather or 'Not recorded'}
{geometry_summary}

Vehicles Involved:
"""
    if vehicle1:
        report += f"- {vehicle1.name}: Mass = {vehicle1.mass} kg, Initial Speed = {initial_v1 * 3.6:.1f} km/h\n"
    if vehicle2:
        report += f"- {vehicle2.name}: Mass = {vehicle2.mass} kg, Initial Speed = {initial_v2 * 3.6:.1f} km/h\n"
    if pedestrian:
        report += f"- {pedestrian.name}: Mass = {pedestrian.mass} kg\n"
    
    recommendations = generate_recommendations(cause, road_condition, lighting, accident_type)
    recommendations_text = "\n".join(f"- {rec}" for rec in recommendations)
    system_summary = generate_systemic_summary(cause, driver_profile, environment, intervention_plan, vehicle1, vehicle2, pedestrian, initial_v1, initial_v2)
    validation_line = validation_text or "Historical validation unavailable."

    baseline_severity = comparison['baseline']['severity'] if comparison else severity
    intervention_severity = comparison['intervention']['severity'] if comparison else severity
    baseline_force = comparison['baseline']['impact_force'] if comparison else impact_force
    intervention_force = comparison['intervention']['impact_force'] if comparison else impact_force
    baseline_risk = comparison['baseline'].get('risk_score') if comparison else risk_score
    intervention_risk = comparison['intervention'].get('risk_score') if comparison else risk_score
    impact_force_reduction = None
    if comparison and baseline_force:
        impact_force_reduction = (baseline_force - intervention_force) / baseline_force * 100
    risk_reduction = None
    if comparison and baseline_risk:
        risk_reduction = baseline_risk - intervention_risk
    baseline_risk_formatted = f"{baseline_risk:.2f}" if baseline_risk is not None else "N/A"
    intervention_risk_formatted = f"{intervention_risk:.2f}" if intervention_risk is not None else "N/A"
    risk_change_text = (
        f"{risk_reduction:.2f} (Baseline {baseline_risk_formatted}, Intervention {intervention_risk_formatted})"
        if risk_reduction is not None and baseline_risk is not None and intervention_risk is not None
        else f"0.00 (Baseline {baseline_risk_formatted}, Intervention {intervention_risk_formatted})"
    )
    force_reduction_text = f"{impact_force_reduction:.2f}%" if impact_force_reduction is not None else "0%"
    intervention_desc = intervention_plan.describe() if intervention_plan else 'None'
    baseline_risk_text = baseline_risk_formatted
    intervention_risk_text = intervention_risk_formatted
    
    report += f"""

Analysis:
This simulation demonstrates how speed, road conditions, and lighting contribute to accidents.
Severity is classified based on impact force: minor (<50kN), moderate (50-150kN), severe (>150kN).
Recommendations:
{recommendations_text}

Systems Interaction Summary:
- {system_summary}

Validation Against Historical Records:
- {validation_line}

Intervention Scenario Comparison:
- Baseline Severity: {baseline_severity}
- Intervention Severity: {intervention_severity}
- Baseline Impact Force: {baseline_force:.2f} N
- Intervention Impact Force: {intervention_force:.2f} N
- Risk Score Change: {risk_change_text}
- Impact Force Reduction: {force_reduction_text}
- Interventions Applied: {intervention_desc}

Machine Learning Prediction:
{ml_prediction if ml_prediction else "Not available"}

"""
    # Create formatted Excel report
    wb = Workbook()
    ws = wb.active
    ws.title = "Accident Report"
    
    # Add title
    ws['A1'] = "Road Accident Simulation Report"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:O1')
    
    # Add headers starting from row 3
    data = {
        'Accident Type': accident_desc,
        'Location': f"{location}, Labo, Camarines Norte",
        'Time of Accident': f"{collision_time:.2f} seconds",
        'Cause': cause,
        'Vehicle 1 Name': vehicle1.name if vehicle1 else '',
        'Vehicle 1 Mass (kg)': vehicle1.mass if vehicle1 else '',
        'Vehicle 1 Initial Speed (km/h)': f"{initial_v1 * 3.6:.1f}" if initial_v1 else '',
        'Vehicle 2 Name': vehicle2.name if vehicle2 else '',
        'Vehicle 2 Mass (kg)': vehicle2.mass if vehicle2 else '',
        'Vehicle 2 Initial Speed (km/h)': f"{initial_v2 * 3.6:.1f}" if initial_v2 else '',
        'Pedestrian Name': pedestrian.name if pedestrian else '',
        'Pedestrian Mass (kg)': pedestrian.mass if pedestrian else '',
        'Road Conditions': road_condition,
        'Lighting Conditions': lighting,
        'Weather': weather or 'Not recorded',
        'Angle of Impact (degrees)': angle_of_impact,
        'Collision Time (s)': f"{collision_time:.2f}",
        'Impact Force (N)': f"{impact_force:.2f}",
        'Severity': severity,
        'Risk Score (0-10)': risk_score if risk_score is not None else baseline_risk,
        'Analysis': 'This simulation demonstrates how speed, road conditions, and lighting contribute to accidents.',
        'Recommendations': recommendations_text,
        'Systems Summary': system_summary,
        'Historical Validation': validation_line,
        'Interventions Applied': intervention_desc,
        'Baseline Severity': baseline_severity,
        'Intervention Severity': intervention_severity,
        'Baseline Impact Force (N)': f"{baseline_force:.2f}",
        'Intervention Impact Force (N)': f"{intervention_force:.2f}",
        'Impact Force Reduction (%)': force_reduction_text,
        'Risk Score Change': risk_change_text,
        'Baseline Risk Score': baseline_risk_text,
        'Intervention Risk Score': intervention_risk_text
    }
    
    headers = list(data.keys())
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
        ws.cell(row=3, column=col).font = Font(bold=True)
        ws.cell(row=3, column=col).fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold color
        ws.cell(row=3, column=col).alignment = Alignment(horizontal='center')
    
    # Add data
    for col, value in enumerate(data.values(), 1):
        ws.cell(row=4, column=col, value=value)
        ws.cell(row=4, column=col).alignment = Alignment(horizontal='center')
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for row in range(1, 5):  # Check rows 1 to 4
            cell = ws.cell(row=row, column=col_num)
            if cell.value and not isinstance(cell, MergedCell):
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 30)  # Cap at 30 for readability
        ws.column_dimensions[column_letter].width = adjusted_width
    
    try:
        wb.save('accident_report.xlsx')
        print("Formatted Excel report saved as 'accident_report.xlsx'")
    except PermissionError:
        print("Could not save Excel report due to permission error. Report printed above.")
    
    return report


def create_labo_roads_geojson(force_regenerate: bool = False) -> Path | None:
    """Ensure a clipped Labo road network GeoJSON exists and return its path."""

    base_dir = Path(__file__).resolve().parent if '__file__' in globals() else Path(os.getcwd())
    gis_dir = base_dir / "GIS"
    output_geojson_path = gis_dir / "labo_roads.geojson"

    # If the GeoJSON already exists and regeneration not forced, check if
    # the barangay names are already assigned. If not, update the file.
    if output_geojson_path.exists() and not force_regenerate:
        try:
            roads = gpd.read_file(output_geojson_path)
            # If 'barangay' column is present we can reuse
            if 'barangay' in roads.columns and not roads['barangay'].isnull().all():
                print(f"Found existing {output_geojson_path} with barangay names. Reusing it for the map overlay.")
                return output_geojson_path
            else:
                print(f"Found existing {output_geojson_path} but barangay names are missing; updating file.")
                # continue on to regenerate/annotate
        except Exception:
            # Any issue when reading the existing file will trigger a regeneration
            print("Existing GeoJSON found but could not be read; regenerating.")

    gis_dir.mkdir(parents=True, exist_ok=True)

    def pick_source(candidates: list[Path]) -> Path | None:
        for candidate in candidates:
            if candidate.exists():
                return candidate
        return None

    road_source = pick_source([
        gis_dir / "highway.shp",
        gis_dir / "Labo_Roads.shp",
        gis_dir / "labo_roads.geojson.shp"
    ])

    boundary_source = pick_source([
        gis_dir / "Labo_Boundary.shp",
        gis_dir / "Labo_Boundary.geojson",
        gis_dir / "Labo_Boundary.shp.gpkg"
    ])

    if road_source is None:
        print("No road network file found in the GIS directory. Expected 'highway.shp' or a Labo-specific layer.")
        return None

    print(f"Reading road network from {road_source}...")
    try:
        roads = gpd.read_file(road_source)
    except Exception as exc:
        print(f"Error reading road network layer: {exc}")
        return None

    clip_required = boundary_source is not None and road_source.name.lower() not in {"labo_roads.geojson.shp", "labo_roads.shp"}

    if clip_required:
        print(f"Reading Labo boundary from {boundary_source}...")
        try:
            boundary = gpd.read_file(boundary_source)
        except Exception as exc:
            print(f"Error reading boundary layer: {exc}")
            return None

        print("Clipping roads to the Labo boundary...")
        try:
            roads = gpd.clip(roads, boundary)
        except Exception as exc:
            print(f"Error during clipping: {exc}")
            return None
    else:
        if boundary_source is None:
            print("Boundary layer not found; exporting the available road layer as-is.")
        else:
            print("Detected an existing Labo-only road layer; skipping additional clipping.")

    if roads.empty:
        print("No road geometries found to export. GeoJSON will not be created.")
        return None

    # Optionally add barangay names to properties
    try:
        roads = _assign_barangay_names(roads)
    except Exception as exc:
        print(f"Warning: Could not assign barangay names to roads: {exc}")

    print(f"Saving Labo road network to {output_geojson_path}...")
    try:
        roads.to_file(output_geojson_path, driver="GeoJSON")
    except Exception as exc:
        print(f"Error saving GeoJSON: {exc}")
        return None

    print("Success! labo_roads.geojson is ready.")
    return output_geojson_path


def create_map_visualization(
    accident_location_or_sim_data: Any = None,
    accident_location: Any | None = None,
    accident_type: Optional[str] = None,
    severity: Optional[str] = None,
    entities: Optional[List[str]] = None,
    timestamp: Optional[datetime] = None,
    animate_vehicles: bool = True,
    vehicle_roads_limit: int = 50,
    vehicles_per_road: int = 1,
    target_fps: int = 30,
    **kwargs
) -> str:
    """
    Creates a map visualization showing the accident location and relevant information.
    Returns HTML content for embedding in Streamlit.
    
    Parameters:
    - target_fps: int - target frames per second for vehicle animation (<= 60). Higher values increase HTML payload and CPU usage.
    """
    print("\nStarting map visualization process...")

    # Backwards-compatible wrapper: support two calling styles
    # 1) (accident_location: (lat, lon), accident_type, severity, entities, timestamp, ...)
    # 2) (sim_data, vehicle1_name=..., vehicle2_name=..., pedestrian_name=..., location=<loc name>, labo_geojson_path=..., collision_time=...)
    # Support both calling styles:
    # - Positional (accident_location_or_sim_data) may be a sim_data DataFrame or a (lat, lon) tuple
    # - Keyword `accident_location` may be used by callers that prefer named parameter
    accident_lat, accident_lon = None, None

    # If caller passed `accident_location` as a keyword and no positional first argument,
    # use that value as the main first param so the rest of the wrapper will work.
    if accident_location_or_sim_data is None and accident_location is not None:
        accident_location_or_sim_data = accident_location
    labo_geojson_path = kwargs.get('labo_geojson_path')
    if hasattr(accident_location_or_sim_data, 'columns'):
        # Called with sim_data
        sim_data = accident_location_or_sim_data
        # extract names from kwargs
        vehicle1_name = kwargs.get('vehicle1_name')
        vehicle2_name = kwargs.get('vehicle2_name')
        pedestrian_name = kwargs.get('pedestrian_name')
        location_name = kwargs.get('location')
        collision_time = kwargs.get('collision_time')
        # Resolve location name to lat/lon using LABO_COORDINATES and fallback grid
        location_coords = {}
        base_lat, base_lon = 14.156, 122.83
        delta_lat = 0.0022
        delta_lon = 0.0032
        # seed with constants mapping
        if isinstance(LABO_COORDINATES, dict):
            for k, v in LABO_COORDINATES.items():
                try:
                    location_coords[k] = [float(v[0]), float(v[1])]
                except Exception:
                    continue
        # fill the rest of barangays with a deterministic grid placement
        per_row = 8
        for i, name in enumerate(LABO_BARANGAYS):
            if name in location_coords:
                continue
            row = i // per_row
            col = i % per_row
            lat = base_lat + (row - 3) * delta_lat
            lon = base_lon + (col - (per_row // 2)) * delta_lon
            location_coords[name] = [round(lat, 6), round(lon, 6)]
        if location_name and location_name in location_coords:
            accident_lat, accident_lon = location_coords.get(location_name)
        else:
            # Default to center of Labo if location can't be resolved
            accident_lat, accident_lon = 14.156, 122.83
        # Use provided entities list or construct from names
        if entities is None:
            entities = [n for n in (vehicle1_name, vehicle2_name, pedestrian_name) if n]
        # accident_type/severity/timestamp may be provided; default to values if missing
        accident_type = accident_type or 'unknown'
        severity = severity or 'unknown'
        timestamp = timestamp or datetime.now()
    else:
        # Old style: first arg is a tuple (lat, lon)
        accident_location = accident_location_or_sim_data
        try:
            accident_lat, accident_lon = accident_location
        except Exception:
            accident_lat, accident_lon = (14.156, 122.83)
        # Entities default
        entities = entities or []
        # Ensure timestamp default for consistency
        timestamp = timestamp or datetime.now()

    # Create a base map centered on the accident location
    m = folium.Map(location=[accident_lat, accident_lon], zoom_start=15)

    # Add accident location marker with animated pulse
    accident_marker_css = """
    <style>
    .accident-marker {
        position: relative;
        width: 34px;
        height: 34px;
    }
    .accident-marker-core {
        position: absolute;
        top: 50%;
        left: 50%;
        width: 18px;
        height: 18px;
        transform: translate(-50%, -50%);
        background: #d9534f;
        border-radius: 50%;
        color: #ffffff;
        display: flex;
        align-items: center;
        justify-content: center;
        box-shadow: 0 0 10px rgba(217, 83, 79, 0.7);
        animation: accident-core-pulse 1.6s infinite;
    }
    .accident-marker-core i {
        font-size: 12px;
    }
    .accident-marker-wave {
        position: absolute;
        top: 50%;
        left: 50%;
        width: 34px;
        height: 34px;
        transform: translate(-50%, -50%);
        border-radius: 50%;
        border: 2px solid rgba(217, 83, 79, 0.7);
        opacity: 0;
        animation: accident-wave 2.4s infinite;
    }
    .accident-marker-wave.wave-delay {
        animation-delay: 1.2s;
    }
    @keyframes accident-wave {
        0% {
            opacity: 0.7;
            transform: translate(-50%, -50%) scale(0.3);
        }
        60% {
            opacity: 0;
            transform: translate(-50%, -50%) scale(1.6);
        }
        100% {
            opacity: 0;
            transform: translate(-50%, -50%) scale(1.6);
        }
    }
    @keyframes accident-core-pulse {
        0%, 100% {
            transform: translate(-50%, -50%) scale(1);
            box-shadow: 0 0 10px rgba(217, 83, 79, 0.7);
        }
        50% {
            transform: translate(-50%, -50%) scale(1.15);
            box-shadow: 0 0 18px rgba(217, 83, 79, 0.9);
        }
    }
    </style>
    """
    m.get_root().header.add_child(folium.Element(accident_marker_css))

    accident_marker_html = """
    <div class="accident-marker">
        <div class="accident-marker-wave"></div>
        <div class="accident-marker-wave wave-delay"></div>
        <div class="accident-marker-core"><i class="fa fa-exclamation"></i></div>
    </div>
    """

    # Create popup content for accident marker
    popup_content = f"""
    <b>Accident Details</b><br>
    Type: {accident_type.replace('_', ' ').title()}<br>
    Severity: {severity.capitalize()}<br>
    Time: {timestamp.strftime('%Y-%m-%d %H:%M:%S')}<br>
    Entities: {', '.join(entities) if entities else 'None'}
    """

    folium.Marker(
        location=[accident_lat, accident_lon],
        popup=popup_content,
        icon=folium.DivIcon(html=accident_marker_html, icon_size=(34, 34), icon_anchor=(17, 17), class_name='')
    ).add_to(m)

    # Add Labo road network overlay if available
    if labo_geojson_path:
        labo_geojson_path = Path(labo_geojson_path)
    else:
        labo_geojson_path = Path(__file__).resolve().parent / "GIS" / "labo_roads.geojson"
    if labo_geojson_path.exists():
        try:
            roads_gdf = gpd.read_file(labo_geojson_path)
            if not roads_gdf.empty:
                roads_geojson_data = json.loads(roads_gdf.to_json())

                from folium import plugins

                def road_style(_):
                    return {"color": "#ff7f0e", "weight": 3, "opacity": 0.85}

                # Add the static GeoJson layer (for selection and base style)
                from folium.features import GeoJsonTooltip
                tooltip = GeoJsonTooltip(fields=['barangay'], aliases=['Barangay:'], localize=True)
                folium.GeoJson(
                    roads_geojson_data,
                    name="Labo Road Network",
                    style_function=road_style,
                    tooltip=tooltip
                ).add_to(m)

                # Also add animated moving traffic using folium.plugins.AntPath for each LineString
                traf_fg = folium.FeatureGroup(name="Moving Traffic Flow")
                for feature in roads_geojson_data.get('features', []):
                    geom = feature.get('geometry', {})
                    props = feature.get('properties', {})
                    if geom.get('type') == 'LineString':
                        # coordinates are [lng, lat]; AntPath expects list of [lat, lng]
                        coords = [[c[1], c[0]] for c in geom.get('coordinates', [])]
                        if len(coords) >= 2:
                            # Create AntPath with color gradient to indicate movement
                            ant = plugins.AntPath(
                                locations=coords,
                                color='#b21f2d',
                                pulse_color='#fffd75',
                                weight=4,
                                delay=2000,
                                opacity=0.9
                            )
                            traf_fg.add_child(ant)

                m.add_child(traf_fg)
                folium.LayerControl().add_to(m)
                print("Added Labo road network overlay")
                # Add TimestampedGeoJson vehicle markers to follow road geometry
                try:
                    from folium.plugins import TimestampedGeoJson
                    # Use a subset of roads to generate vehicles to keep HTML size reasonable
                    if animate_vehicles:
                        parsed_roads = parse_labo_roads_geojson()
                        # Choose frame rate: cap to reasonable values (<= 60 fps)
                        fps = int(kwargs.get('target_fps', target_fps) or target_fps)
                        if fps <= 0:
                            fps = 1
                        if fps > 60:
                            print(f"Requested fps {fps} exceeds 60 -- capping to 60 for browser performance")
                            fps = 60
                        # Limit number of roads to animate to the configured number for performance
                        # Create a deterministic RNG for the location if available
                        loc_rng = rng_for_location(location_name) if ('location_name' in locals() and location_name) else None
                        vehicle_geojson = _generate_timestamped_vehicle_geojson(parsed_roads[:vehicle_roads_limit], vehicles_per_road=vehicles_per_road, fps=fps, rng=loc_rng)
                        feature_count = len(vehicle_geojson.get('features', [])) if vehicle_geojson else 0
                        print(f"Timestamped vehicle features generated: {feature_count}")
                        # Warn users about potential performance implications
                        if fps >= 30 and (vehicle_roads_limit > 20 or vehicles_per_road > 2 or feature_count > 200):
                            print("Warning: Generating high-frame-rate animations (>=30fps) for many roads/vehicles can produce very large HTML payloads and may degrade browser performance. Consider reducing fps, vehicle_roads_limit, or vehicles_per_road.")
                        # Choose frame rate: cap to reasonable values (<= 60 fps)
                        fps = int(kwargs.get('target_fps', target_fps) or target_fps)
                        if fps <= 0:
                            fps = 1
                        if fps > 60:
                            print(f"Requested fps {fps} exceeds 60 -- capping to 60 for browser performance")
                            fps = 60

                        TimestampedGeoJson(
                            vehicle_geojson,
                            # Period string describes the time step between frames. Set to fraction of a second per frame.
                            period=f'PT{1.0/fps:.6f}S',
                            add_last_point=True,
                            auto_play=True,
                            loop=True,
                            max_speed=1,
                            loop_button=True,
                            date_options='YYYY/MM/DD HH:mm:ss',
                            time_slider_drag_update=True
                        ).add_to(m)
                        print("Added timestamped vehicle layer to map")
                except Exception as e:
                    print(f"Could not add timestamped vehicle markers: {e}")
        except Exception as e:
            print(f"Could not load Labo roads: {e}")

    # Add historical accident data markers
    try:
        historical_data = pd.read_csv('accident_data.csv')
        location_counts = historical_data.groupby('location').size().reset_index(name='accident_count')

        # Accurate coordinates for Labo locations
        location_coords = {
            'Bayabas': [14.158, 122.825],
            'Main Highway': [14.156, 122.83],
            'Barangay 1': [14.152, 122.835],
            'Barangay 2': [14.154, 122.828],
            'Barangay 3': [14.160, 122.832],
            'Barangay 4': [14.158, 122.838],
            'Barangay 5': [14.162, 122.826],
        }

        for _, row in location_counts.iterrows():
            loc_name = row['location']
            count = row['accident_count']

            if loc_name in location_coords:
                marker_lat, marker_lon = location_coords[loc_name]
            else:
                continue

            # Color based on accident count
            if count >= 5:
                color = 'red'
            elif count >= 3:
                color = 'orange'
            else:
                color = 'blue'

            folium.Marker(
                location=[marker_lat, marker_lon],
                popup=f"<b>{loc_name}</b><br>Historical Accidents: {count}",
                icon=folium.Icon(color=color, icon='exclamation-triangle', prefix='fa')
            ).add_to(m)

        print(f"Added {len(location_counts)} historical accident markers")
    except Exception as e:
        print(f"Could not load historical accident data: {e}")

    # Add moving traffic annotations
    roads = parse_labo_roads_geojson()
    # We added AntPath animated polylines above (via folium.plugins.AntPath) tied to the road coordinates.
    print("Added animated AntPath moving traffic overlays tied to the Labo road GeoJSON.")

    # Add legend
    from branca.element import MacroElement
    from jinja2 import Template

    legend_html = """
    {% macro html(this, kwargs) %}
    <div style="position: fixed; bottom: 20px; left: 20px; z-index: 9999; background-color: white; border: 1px solid #ccc; border-radius: 6px; box-shadow: 0 2px 6px rgba(0,0,0,0.25); padding: 12px 14px; font-size: 13px; line-height: 1.5; min-width: 250px;">
        <div style="font-weight: 600; margin-bottom: 8px;">Map Legend</div>
        <div style="display: flex; align-items: center; margin-bottom: 6px;">
            <span style="display: inline-block; width: 26px; height: 4px; background-color: #ff7f0e; margin-right: 8px;"></span>
            <span>Labo road network</span>
        </div>
        <div style="display: flex; align-items: center; margin-bottom: 6px;">
            <i class="fa fa-exclamation" style="color: #d9534f; margin-right: 8px;"></i>
            <span>Current accident location</span>
        </div>
        <div style="display: flex; align-items: center; margin-bottom: 6px;">
            <div style="width: 12px; height: 12px; border-radius: 50%; background: linear-gradient(90deg, #ff4444, #4444ff, #44ff44); margin-right: 8px; animation: pulse 2s infinite;"></div>
            <span>Moving traffic flow</span>
        </div>
        <div style="display: flex; align-items: center; margin-bottom: 6px;">
            <i class="fa fa-exclamation-triangle" style="color: #d9534f; margin-right: 8px;"></i>
            <span>High accident density (≥5)</span>
        </div>
        <div style="display: flex; align-items: center; margin-bottom: 6px;">
            <i class="fa fa-exclamation-triangle" style="color: #f0ad4e; margin-right: 8px;"></i>
            <span>Moderate accident density (3-4)</span>
        </div>
        <div style="display: flex; align-items: center;">
            <i class="fa fa-exclamation-triangle" style="color: #0275d8; margin-right: 8px;"></i>
            <span>Low accident density (&lt;3)</span>
        </div>
    </div>
    {% raw %}
    <style>
    @keyframes pulse {
        0%, 100% { opacity: 1; }
        50% { opacity: 0.5; }
    }
    </style>
    {% endraw %}
    {% endmacro %}
    """

    legend = MacroElement()
    legend._template = Template(legend_html)
    m.get_root().add_child(legend)

    # Add favicon to avoid 404 requests
    favicon_link = '<link rel="icon" href="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMBApZfo7kAAAAASUVORK5CYII=" />'
    m.get_root().header.add_child(folium.Element(favicon_link))

    # Return the HTML content instead of saving to file
    import tempfile
    import os

    # Create a temporary file to save the map
    with tempfile.NamedTemporaryFile(mode='w+', suffix='.html', delete=False, encoding='utf-8') as temp_file:
        temp_filename = temp_file.name
        m.save(temp_filename)

    # Read the HTML content from the temporary file
    try:
        with open(temp_filename, 'r', encoding='utf-8') as f:
            html_string = f.read()
    finally:
        # Clean up the temporary file
        os.unlink(temp_filename)

    print("Map visualization generated successfully for Streamlit integration.")
    return html_string
    
def _haversine_meters(lat1, lon1, lat2, lon2):
    """Return distance between two lat/lon points in meters (haversine)."""
    from math import radians, sin, cos, asin, sqrt
    R = 6371000  # Earth radius in meters
    phi1, phi2 = radians(lat1), radians(lat2)
    dphi = radians(lat2 - lat1)
    dlambda = radians(lon2 - lon1)
    a = sin(dphi/2)**2 + cos(phi1) * cos(phi2) * sin(dlambda/2)**2
    c = 2 * asin(min(1, sqrt(a)))
    return R * c


def _assign_barangay_names(roads_gdf, barangay_centers: dict | None = None, max_radius_m: float = 3000.0):
    """Assign a barangay name to each geometry in a GeoDataFrame based on nearest
    center points. Adds a 'barangay' column to the GeoDataFrame.

    The function uses the centroid of each geometry to find the nearest
    barangay centre within max_radius_m. If none are within the radius, it
    assigns 'Unknown'.
    """
    if barangay_centers is None:
        # Build default centers using the LABO_BARANGAYS list. Where available, prefer
        # coordinates from LABO_COORDINATES; otherwise create a small grid around the
        # Labo town center (14.156, 122.83) for consistent but approximate positions.
        base_lat, base_lon = 14.156, 122.83
        barangay_centers = {}
        # Use any explicit coords in LABO_COORDINATES
        known_coords = LABO_COORDINATES if 'LABO_COORDINATES' in globals() else {}
        delta_lat = 0.0022
        delta_lon = 0.0032
        for i, name in enumerate(LABO_BARANGAYS):
            if name in known_coords:
                barangay_centers[name] = tuple(known_coords[name])
                continue
            # grid placement: rows of 8
            per_row = 8
            row = i // per_row
            col = i % per_row
            lat = base_lat + (row - 3) * delta_lat  # center rows around base
            lon = base_lon + (col - (per_row // 2)) * delta_lon
            barangay_centers[name] = (round(lat, 6), round(lon, 6))

    # Build a list for iteration
    centers_list = [(name, coords[0], coords[1]) for name, coords in barangay_centers.items()]

    def nearest_barangay(lat, lon):
        best_name = 'Unknown'
        best_dist = float('inf')
        for name, c_lat, c_lon in centers_list:
            d = _haversine_meters(lat, lon, c_lat, c_lon)
            if d < best_dist:
                best_dist = d
                best_name = name
        if best_dist <= max_radius_m:
            return best_name
        else:
            return 'Unknown'

    # Use centroid of geometry (lat=centroid.y, lon=centroid.x)
    try:
        centroids = roads_gdf.geometry.centroid
    except Exception:
        # Try converting to geometry column if necessary
        centroids = roads_gdf.apply(lambda r: r.geometry.centroid if hasattr(r, 'geometry') else None, axis=1)

    barangays = []
    for c in centroids:
        if c is None:
            barangays.append('Unknown')
            continue
        lat = float(c.y)
        lon = float(c.x)
        barangays.append(nearest_barangay(lat, lon))

    roads_gdf = roads_gdf.copy()
    roads_gdf['barangay'] = barangays
    return roads_gdf


def _generate_timestamped_vehicle_geojson(roads, start_time=None, vehicles_per_road=1, fps=30, rng: Optional[random.Random] = None):
    """
    Generate a GeoJSON FeatureCollection for TimestampedGeoJson plugin where each feature
    represents a vehicle moving along a LineString path with matching 'times' array.
    `roads` should be a list of roads as returned by `parse_labo_roads_geojson` (lat, lng points).
    """
    import copy
    from datetime import datetime, timedelta
    if start_time is None:
        # Use UTC now as the animation baseline
        start_time = datetime.utcnow()

    features = []

    r = rng or random
    for road_idx, road in enumerate(roads):
        coords_latlng = road.get('coordinates', [])
        if len(coords_latlng) < 2:
            continue

        # We'll place a number of vehicles per road; compute once
        for v_idx in range(vehicles_per_road):
            # Choose a vehicle speed (m/s) - vary per vehicle
            # realistic range: cars 8-16 m/s (~30-60 km/h), bikes slower
            speed_mps = r.uniform(8.0, 16.0)

            # Choose a start offset on the road [0..1]
            start_offset = r.uniform(0, 0.5)

            # Build the path coordinates as GeoJSON ([lon, lat]) and compute distances
            geojson_coords = []
            distances = []  # meters between consecutive points
            for i, (lat, lon) in enumerate(coords_latlng):
                geojson_coords.append([lon, lat])
                if i > 0:
                    prev_lat, prev_lon = coords_latlng[i-1]
                    d = _haversine_meters(prev_lat, prev_lon, lat, lon)
                    distances.append(d)

            total_length = sum(distances) if distances else 0.0
            if total_length <= 0:
                continue

            # compute cumulative distances along points
            cum_dists = [0.0]
            running = 0.0
            for d in distances:
                running += d
                cum_dists.append(running)

            # Determine time for the whole road (total length / speed)
            total_time = total_length / speed_mps
            base_time = start_time + timedelta(seconds=r.uniform(0, 10))

                # Create a dense sample of times at the requested fps
            fps = int(fps or 60)
            if fps <= 0:
                fps = 1
            if fps > 60:
                fps = 60
            dt = 1.0 / fps
            # times for the whole path [0 .. total_time] sampled every dt
            time_offsets = [i * dt for i in range(int(total_time / dt) + 1)]

            # Interpolate coordinates for each time offset so that we have one coordinate per timestamp
            path_coords = []
            path_times = []
            for offset in time_offsets:
                dist_at_t = min(offset * speed_mps, total_length)
                # Find segment
                seg_idx = 0
                while seg_idx < len(cum_dists) - 1 and cum_dists[seg_idx+1] < dist_at_t:
                    seg_idx += 1
                # Interpolate between seg_idx and seg_idx+1
                if seg_idx >= len(cum_dists) - 1:
                    # we're at or beyond the last point
                    lat, lon = coords_latlng[-1]
                else:
                    dist0 = cum_dists[seg_idx]
                    dist1 = cum_dists[seg_idx+1]
                    # avoid division by zero
                    if dist1 - dist0 == 0:
                        frac = 0
                    else:
                        frac = (dist_at_t - dist0) / (dist1 - dist0)
                    lat0, lon0 = coords_latlng[seg_idx]
                    lat1, lon1 = coords_latlng[seg_idx+1]
                    lat = lat0 + frac * (lat1 - lat0)
                    lon = lon0 + frac * (lon1 - lon0)
                path_coords.append([lon, lat])
                path_times.append((base_time + timedelta(seconds=offset)).isoformat() + 'Z')

            # If we want vehicles to start mid-segment, we can rotate the coords/times arrays
            # to apply start_offset (fraction of total length)
            offset_m = start_offset * total_length
            # find index on the densified path to rotate
            dense_cum_dists = [0.0]
            running = 0.0
            for i in range(1, len(path_coords)):
                prev = path_coords[i-1]
                curr = path_coords[i]
                d = _haversine_meters(prev[1], prev[0], curr[1], curr[0])
                running += d
                dense_cum_dists.append(running)

            rot_idx = 0
            for i, d in enumerate(dense_cum_dists):
                if d >= offset_m:
                    rot_idx = i
                    break

            # rotate so vehicle starts from rot_idx and append earlier coords with times increased by total_time
            if rot_idx > 0:
                path_coords = path_coords[rot_idx:] + path_coords[:rot_idx]
                # Adjust times to keep monotonicity by pushing the wrapped times to the end
                wrapped_times = []
                for t in path_times[:rot_idx]:
                    wrapped_times.append((datetime.fromisoformat(t.replace('Z','')) + timedelta(seconds=total_time)).isoformat() + 'Z')
                path_times = path_times[rot_idx:] + wrapped_times

            # Keep times length equal to coordinates length
            # Ensure equal lengths for coords and times (they should be equal by construction)
            if len(path_times) != len(path_coords):
                minlen = min(len(path_times), len(path_coords))
                path_coords = path_coords[:minlen]
                path_times = path_times[:minlen]

            # assign a color per vehicle type (round-robin)
            colors = ['#ff4444', '#4444ff', '#44ff44']
            color = colors[(road_idx + v_idx) % len(colors)]

            feature = {
                'type': 'Feature',
                'geometry': {'type': 'LineString', 'coordinates': path_coords},
                'properties': {
                    'times': path_times,
                    'popup': f'Vehicle {road_idx}-{v_idx}',
                    'style': {'color': color, 'weight': 4},
                    'icon': 'circle',
                    'iconstyle': {
                        'fillColor': color,
                        'fillOpacity': 0.9,
                        'stroke': True,
                        'radius': 6
                    }
                }
            }

            features.append(feature)

    return {
        'type': 'FeatureCollection',
        'features': features
    }
def generate_pdf_report(report_text: str, filename: str = "accident_report.pdf"):
    """Generate a PDF report from the text report."""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer

        doc = SimpleDocTemplate(filename, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []

        # Title
        title = Paragraph("Road Accident Simulation Report", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 12))

        # Report content
        for line in report_text.split('\n'):
            if line.strip():
                p = Paragraph(line, styles['Normal'])
                story.append(p)
            else:
                story.append(Spacer(1, 6))

        doc.build(story)
        print(f"PDF report generated: {filename}")
        return filename
    except ImportError:
        print("reportlab not installed. Install with: pip install reportlab")
        return None